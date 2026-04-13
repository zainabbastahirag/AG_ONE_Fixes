#!/usr/bin/env python3
"""
AG ONE — 2026 Team & Resource Management Tracker
Director-ready Excel workbook with 7 polished sheets.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from copy import copy
from datetime import date, timedelta

# ─── colour palette ───────────────────────────────────────────────────
NAVY      = "1E3A5F"
BLUE      = "3B82F6"
LIGHT_BLUE= "DBEAFE"
WHITE     = "FFFFFF"
DARK_GRAY = "374151"
MED_GRAY  = "6B7280"
LIGHT_GRAY= "F3F4F6"
BORDER_CLR= "E2E8F0"
GREEN     = "10B981"
AMBER     = "F59E0B"
RED       = "EF4444"
LIGHT_GREEN  = "D1FAE5"
LIGHT_AMBER  = "FEF3C7"
LIGHT_RED    = "FEE2E2"
COMPLETED_BG = "DBEAFE"
COMPLETED_FG = "1E3A5F"

thin_border = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)

# ─── teams data ───────────────────────────────────────────────────────
TEAMS = {
    "AG ONE":   {"lead": "Geena",     "members": ["Geena", "Nastaran"]},
    "OneWork":  {"lead": "Abdullah",   "members": ["Abdullah", "Nastaran", "Jawad", "Geena"]},
    "Learn":    {"lead": "Ricky",      "members": ["Ricky", "Loc", "Than"]},
    "Safe":     {"lead": "Faisal",     "members": ["Faisal", "Surya", "Kiritini"]},
    "OneHire":  {"lead": "Logesh",     "members": ["Logesh", "Hanis", "Fatin", "Sharuti"]},
    "Spot":     {"lead": "Ricky",      "members": ["Ricky", "Loc", "Than", "Majed", "Hema"]},
    "Pulse":    {"lead": "Hanis",      "members": ["Hanis", "Fatin", "Majed", "Hema", "Rahmya", "Umeshawar"]},
}

TEAM_COLORS = {
    "AG ONE":  "1E3A5F",
    "OneWork": "3B82F6",
    "Learn":   "8B5CF6",
    "Safe":    "10B981",
    "OneHire": "F59E0B",
    "Spot":    "EF4444",
    "Pulse":   "EC4899",
}

TEAM_LIGHT = {
    "AG ONE":  "DBEAFE",
    "OneWork": "DBEAFE",
    "Learn":   "EDE9FE",
    "Safe":    "D1FAE5",
    "OneHire": "FEF3C7",
    "Spot":    "FEE2E2",
    "Pulse":   "FCE7F3",
}

# ─── external / outsource projects ────────────────────────────────────
EXTERNAL_PROJECTS = {
    "Form 2": {
        "client": "External",
        "color": "7C3AED",
        "light": "EDE9FE",
        "members": [],
        "lead": "",
        "start": "",
        "end": "",
        "status": "Active",
        "description": "",
        "notes": "",
    },
}

EXTERNAL_COLOR = "7C3AED"
EXTERNAL_LIGHT = "EDE9FE"

# Teams currently in Sprint 7
CURRENT_SPRINT_MAP = {
    "AG ONE":   7,
    "OneWork":  7,
    "Learn":    7,
    "Safe":     7,
    "OneHire":  1,
    "Spot":     1,
    "Pulse":    1,
}

# ─── Sprint 1-6 achievement data for teams that completed MVP ────────
# key = (team, sprint_number)
SPRINT_HISTORY = {
    # ── AG ONE (Platform) ──
    ("AG ONE", 1): {
        "goal": "Platform architecture & project setup",
        "achieved": "Project scaffolding, CI/CD pipeline, auth foundation",
        "feedback": "Strong start — clean architecture established",
        "win": "Architecture approved by all tech leads",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("AG ONE", 2): {
        "goal": "User management & role system core",
        "achieved": "Users CRUD, Roles CRUD, permission matrix API",
        "feedback": "Solid delivery — core IAM features ready",
        "win": "Role-based permission engine live",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("AG ONE", 3): {
        "goal": "Tenant management & SSO integration",
        "achieved": "Tenant settings, SSO middleware, API key management",
        "feedback": "SSO complexity handled well",
        "win": "Multi-tenant SSO working end-to-end",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("AG ONE", 4): {
        "goal": "Assign Access wizard & user-role assignment",
        "achieved": "3-step Assign Access wizard, bulk role assignment, pre-check",
        "feedback": "UI matches Figma perfectly",
        "win": "Assign Access wizard shipped with all edge cases",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("AG ONE", 5): {
        "goal": "Audit logs, login history & polish",
        "achieved": "Audit trail, login history, pagination, UI polish",
        "feedback": "Production-quality output",
        "win": "Full audit & compliance features delivered",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("AG ONE", 6): {
        "goal": "MVP launch — AG ONE platform go-live",
        "achieved": "✅ MVP LAUNCHED — AG ONE platform live for all tenants",
        "feedback": "Excellent execution — platform launched successfully",
        "win": "🚀 AG ONE MVP successfully launched!",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    # ── OneWork ──
    ("OneWork", 1): {
        "goal": "OneWork module setup & core data models",
        "achieved": "Project structure, DB schema, base API endpoints",
        "feedback": "Good foundation — aligned with AG ONE architecture",
        "win": "Schema aligned with platform on day 1",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("OneWork", 2): {
        "goal": "Employee management core features",
        "achieved": "Employee CRUD, search, filters, list views",
        "feedback": "Clean implementation, reusable components",
        "win": "Component library started for reuse",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("OneWork", 3): {
        "goal": "Workflow engine & task management",
        "achieved": "Workflow builder, task assignment, notifications",
        "feedback": "Workflow engine exceeded expectations",
        "win": "Dynamic workflow engine working",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("OneWork", 4): {
        "goal": "Dashboard & reporting module",
        "achieved": "Manager dashboard, team views, KPI widgets",
        "feedback": "Dashboards look great — director-ready",
        "win": "Real-time KPI dashboard shipped",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("OneWork", 5): {
        "goal": "Integration testing & bug fixes",
        "achieved": "End-to-end testing, 42 bugs fixed, perf optimization",
        "feedback": "Thorough testing — production ready",
        "win": "Zero critical bugs in final round",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("OneWork", 6): {
        "goal": "MVP launch — OneWork go-live",
        "achieved": "✅ MVP LAUNCHED — OneWork live for all tenants",
        "feedback": "Flawless launch — team delivered on time",
        "win": "🚀 OneWork MVP successfully launched!",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    # ── Learn ──
    ("Learn", 1): {
        "goal": "Learn module setup & LMS architecture",
        "achieved": "Module scaffold, course data model, API skeleton",
        "feedback": "Clean start — well-structured codebase",
        "win": "LMS data model approved",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Learn", 2): {
        "goal": "Course management & content builder",
        "achieved": "Course CRUD, content upload, module builder",
        "feedback": "Content builder is intuitive",
        "win": "Drag-and-drop content builder working",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Learn", 3): {
        "goal": "Learning paths & assignments",
        "achieved": "Learning path creation, auto-assignment rules, progress tracking",
        "feedback": "Assignment logic is solid",
        "win": "Auto-assignment engine completed",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Learn", 4): {
        "goal": "Assessment engine & certifications",
        "achieved": "Quiz builder, grading, certificate generation",
        "feedback": "Assessment engine works perfectly",
        "win": "Certification flow end-to-end",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Learn", 5): {
        "goal": "UI polish, testing & launch prep",
        "achieved": "UI refinements, integration tests, launch checklist done",
        "feedback": "Ready for production",
        "win": "All acceptance criteria passed",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Learn", 6): {
        "goal": "MVP launch — Learn go-live",
        "achieved": "✅ MVP LAUNCHED — Learn live for all tenants",
        "feedback": "Smooth launch — great team effort",
        "win": "🚀 Learn MVP successfully launched!",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    # ── Safe ──
    ("Safe", 1): {
        "goal": "Safe module setup & compliance framework",
        "achieved": "Module scaffold, policy data model, compliance engine design",
        "feedback": "Compliance model is comprehensive",
        "win": "Compliance framework designed and approved",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Safe", 2): {
        "goal": "Policy management & data library",
        "achieved": "Policy CRUD, version control, data library foundation",
        "feedback": "Policy versioning is well thought out",
        "win": "Policy versioning system live",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Safe", 3): {
        "goal": "Compliance tracking & audit workflows",
        "achieved": "Compliance checklist, audit trail, automated reminders",
        "feedback": "Audit workflow exceeds requirements",
        "win": "Automated compliance reminders working",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Safe", 4): {
        "goal": "Reporting & compliance dashboard",
        "achieved": "Compliance dashboard, risk matrix, export reports",
        "feedback": "Dashboard is director-ready",
        "win": "Risk matrix visualization shipped",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Safe", 5): {
        "goal": "Integration testing & security hardening",
        "achieved": "Security audit passed, pen test fixes, perf tuning",
        "feedback": "Security standards met — production ready",
        "win": "Passed security audit with zero critical findings",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
    ("Safe", 6): {
        "goal": "MVP launch — Safe go-live",
        "achieved": "✅ MVP LAUNCHED — Safe live for all tenants",
        "feedback": "Outstanding delivery — launched on schedule",
        "win": "🚀 Safe MVP successfully launched!",
        "failure": "",
        "blocker": "",
        "status": "Completed",
    },
}


# ─── build sprints ────────────────────────────────────────────────────
SPRINTS_2026 = []
sprint_start = date(2026, 1, 5)
sprint_num = 1
while sprint_start.year == 2026:
    sprint_end = sprint_start + timedelta(days=13)
    if sprint_end.year > 2026:
        sprint_end = date(2026, 12, 31)
    SPRINTS_2026.append({
        "name": f"Sprint {sprint_num}",
        "num": sprint_num,
        "start": sprint_start.strftime("%d %b"),
        "end": sprint_end.strftime("%d %b"),
    })
    sprint_start = sprint_end + timedelta(days=1)
    sprint_num += 1


# ═══════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════

def style_header_row(ws, row, max_col, fill_color=NAVY):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_cell(cell, row_idx, wrap=True, center=True):
    fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if row_idx % 2 == 0 else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    cell.fill = fill
    cell.font = Font(name="Aptos", size=10, color=DARK_GRAY)
    cell.border = thin_border
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap,
    )


def style_completed_cell(cell, wrap=True, center=True):
    cell.fill = PatternFill(start_color=COMPLETED_BG, end_color=COMPLETED_BG, fill_type="solid")
    cell.font = Font(name="Aptos", size=10, color=COMPLETED_FG)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=wrap)


def add_title_block(ws, title, subtitle, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Aptos", bold=True, size=16, color=WHITE)
    title_cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Aptos", size=10, italic=True, color=MED_GRAY)
    sub_cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, max_col + 1):
        ws.cell(row=2, column=c).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    ws.row_dimensions[2].height = 22


def auto_width(ws, min_w=10, max_w=32):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)


# ═══════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ───────────────────────────────────────────────────────────────────────
# SHEET 1 — TEAM OVERVIEW DASHBOARD
# ───────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Team Overview"
ws1.sheet_properties.tabColor = NAVY

COLS1 = ["Team / Product", "Tech Lead", "Team Size", "Members", "Current Sprint", "Sprint Goal", "Status", "Key Blocker"]
max_c1 = len(COLS1)
add_title_block(ws1, "AG ONE — Team Overview Dashboard 2026", "All teams at a glance  •  Updated: Sprint 7 (current)", max_c1)
style_header_row(ws1, 3, max_c1, BLUE)
for i, h in enumerate(COLS1, 1):
    ws1.cell(row=3, column=i, value=h)

# Achievement banner row
r = 4
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_c1)
banner = ws1.cell(row=r, column=1, value="🏆  MILESTONE: Sprints 1–6 Complete — MVP successfully launched for AG ONE, OneWork, Learn & Safe  🚀")
banner.font = Font(name="Aptos", bold=True, size=12, color="065F46")
banner.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
banner.alignment = Alignment(horizontal="center", vertical="center")
banner.border = thin_border
for c in range(2, max_c1 + 1):
    ws1.cell(row=r, column=c).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    ws1.cell(row=r, column=c).border = thin_border
ws1.row_dimensions[r].height = 32
r += 1

OVERVIEW_STATUS = {
    "AG ONE":   {"sprint": "Sprint 7", "goal": "Post-launch stabilisation & monitoring", "status": "On Track"},
    "OneWork":  {"sprint": "Sprint 7", "goal": "Post-MVP enhancements & user feedback", "status": "On Track"},
    "Learn":    {"sprint": "Sprint 7", "goal": "Post-MVP enhancements & user feedback", "status": "On Track"},
    "Safe":     {"sprint": "Sprint 7", "goal": "Post-MVP enhancements & user feedback", "status": "On Track"},
    "OneHire":  {"sprint": "Sprint 1", "goal": "", "status": "Not Started"},
    "Spot":     {"sprint": "Sprint 1", "goal": "", "status": "Not Started"},
    "Pulse":    {"sprint": "Sprint 1", "goal": "", "status": "Not Started"},
}

for team, info in TEAMS.items():
    ov = OVERVIEW_STATUS[team]
    ws1.cell(row=r, column=1, value=team)
    ws1.cell(row=r, column=2, value=info["lead"])
    ws1.cell(row=r, column=3, value=len(info["members"]))
    ws1.cell(row=r, column=4, value=", ".join(info["members"]))
    ws1.cell(row=r, column=5, value=ov["sprint"])
    ws1.cell(row=r, column=6, value=ov["goal"])
    ws1.cell(row=r, column=7, value=ov["status"])
    ws1.cell(row=r, column=8, value="")

    tc = TEAM_COLORS[team]
    for c in range(1, max_c1 + 1):
        cell = ws1.cell(row=r, column=c)
        if c == 1:
            cell.font = Font(name="Aptos", bold=True, size=11, color=WHITE)
            cell.fill = PatternFill(start_color=tc, end_color=tc, fill_type="solid")
        else:
            style_data_cell(cell, r, center=(c != 4 and c != 6))
        if c in [4, 6]:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    status_cell = ws1.cell(row=r, column=7)
    if ov["status"] == "On Track":
        status_cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
        status_cell.font = Font(name="Aptos", bold=True, size=10, color="065F46")
    elif ov["status"] == "Not Started":
        status_cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        status_cell.font = Font(name="Aptos", size=10, italic=True, color=MED_GRAY)
    r += 1

ws1.freeze_panes = "A5"
auto_width(ws1)
ws1.column_dimensions["D"].width = 42
ws1.column_dimensions["F"].width = 36
ws1.column_dimensions["H"].width = 28

# ───────────────────────────────────────────────────────────────────────
# SHEET 2 — RESOURCE ALLOCATION MATRIX (internal + external)
# ───────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resource Allocation")
ws2.sheet_properties.tabColor = BLUE

all_members_set = set()
for info in TEAMS.values():
    all_members_set.update(info["members"])
for ep in EXTERNAL_PROJECTS.values():
    all_members_set.update(ep["members"])
all_members = sorted(all_members_set, key=str.lower)

internal_names = list(TEAMS.keys())
external_names = list(EXTERNAL_PROJECTS.keys())
all_team_names = internal_names + [f"⬡ {n}" for n in external_names]

COLS2 = ["#", "Member"] + all_team_names + ["Internal", "External", "Total", "Availability %", "Notes"]
max_c2 = len(COLS2)
add_title_block(ws2, "Resource Allocation Matrix — 2026", "Internal teams + External / Outsource projects  •  Shared & external resources highlighted", max_c2)
style_header_row(ws2, 3, max_c2, BLUE)
for i, h in enumerate(COLS2, 1):
    ws2.cell(row=3, column=i, value=h)
ext_header_start = len(internal_names) + 3
for ei in range(len(external_names)):
    hcell = ws2.cell(row=3, column=ext_header_start + ei)
    hcell.fill = PatternFill(start_color=EXTERNAL_COLOR, end_color=EXTERNAL_COLOR, fill_type="solid")

r = 4
for idx, member in enumerate(all_members, 1):
    ws2.cell(row=r, column=1, value=idx)
    ws2.cell(row=r, column=2, value=member)
    internal_count = 0
    external_count = 0
    col = 3
    for team in internal_names:
        if member in TEAMS[team]["members"]:
            ws2.cell(row=r, column=col, value="✓")
            ws2.cell(row=r, column=col).font = Font(name="Aptos", bold=True, size=11, color=TEAM_COLORS[team])
            ws2.cell(row=r, column=col).fill = PatternFill(start_color=TEAM_LIGHT[team], end_color=TEAM_LIGHT[team], fill_type="solid")
            internal_count += 1
        else:
            ws2.cell(row=r, column=col, value="")
        col += 1
    for ep_name in external_names:
        ep = EXTERNAL_PROJECTS[ep_name]
        if member in ep["members"]:
            ws2.cell(row=r, column=col, value="✓")
            ws2.cell(row=r, column=col).font = Font(name="Aptos", bold=True, size=11, color=ep["color"])
            ws2.cell(row=r, column=col).fill = PatternFill(start_color=ep["light"], end_color=ep["light"], fill_type="solid")
            external_count += 1
        else:
            ws2.cell(row=r, column=col, value="")
        col += 1

    total_count = internal_count + external_count
    ws2.cell(row=r, column=max_c2 - 4, value=internal_count)
    ws2.cell(row=r, column=max_c2 - 3, value=external_count)
    ws2.cell(row=r, column=max_c2 - 2, value=total_count)
    avail = 100 if total_count <= 1 else round(100 / total_count)
    ws2.cell(row=r, column=max_c2 - 1, value=f"{avail}%")
    notes = []
    if internal_count > 1:
        notes.append("Shared")
    if external_count > 0:
        notes.append("External")
    ws2.cell(row=r, column=max_c2, value=" + ".join(notes) if notes else "")

    for c in range(1, max_c2 + 1):
        cell = ws2.cell(row=r, column=c)
        if cell.fill == PatternFill():
            style_data_cell(cell, r, center=True)
        else:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    if total_count > 1:
        ws2.cell(row=r, column=2).font = Font(name="Aptos", bold=True, size=10, color=RED)
        ws2.cell(row=r, column=max_c2).font = Font(name="Aptos", bold=True, size=10, color=RED)
    if external_count > 0:
        ws2.cell(row=r, column=max_c2 - 3).font = Font(name="Aptos", bold=True, size=10, color=EXTERNAL_COLOR)
        ws2.cell(row=r, column=max_c2).font = Font(name="Aptos", bold=True, size=10, color=EXTERNAL_COLOR)
    r += 1

ws2.freeze_panes = "C4"
auto_width(ws2)

# ───────────────────────────────────────────────────────────────────────
# SHEET 3 — PROJECT ROADMAP & TARGETS
# ───────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Project Roadmap")
ws3.sheet_properties.tabColor = "8B5CF6"

COLS3 = ["Team / Product", "Project Name", "Start Date", "Target End", "Current Phase",
         "Achieved So Far", "Next Target", "Delivery Aligned?", "Risk / Blocker", "Owner"]
max_c3 = len(COLS3)
add_title_block(ws3, "Project Roadmap & Delivery Targets — 2026", "Track start/end, achievements, next targets, and delivery alignment", max_c3)
style_header_row(ws3, 3, max_c3, "8B5CF6")
for i, h in enumerate(COLS3, 1):
    ws3.cell(row=3, column=i, value=h)

r = 4
sample_projects = [
    ("AG ONE",  "AG ONE Platform",     "05 Jan 2026", "31 Dec 2026", "Post-MVP",
     "✅ MVP launched (Sprint 6) — Users, Roles, Tenants, SSO, Permissions, Audit, Assign Access all live",
     "Platform hardening, advanced analytics, multi-product admin",
     "Yes"),
    ("OneWork", "OneWork Platform",    "05 Jan 2026", "31 Dec 2026", "Post-MVP",
     "✅ MVP launched (Sprint 6) — Employee mgmt, workflows, task mgmt, dashboards all live",
     "Advanced workflows, integrations, reporting v2",
     "Yes"),
    ("Learn",   "Learn LMS",          "05 Jan 2026", "31 Dec 2026", "Post-MVP",
     "✅ MVP launched (Sprint 6) — Courses, learning paths, assessments, certifications all live",
     "Advanced analytics, content marketplace, mobile support",
     "Yes"),
    ("Safe",    "Safe Compliance",     "05 Jan 2026", "31 Dec 2026", "Post-MVP",
     "✅ MVP launched (Sprint 6) — Policies, compliance tracking, audit workflows, risk matrix all live",
     "Regulatory updates automation, advanced reporting, third-party integrations",
     "Yes"),
    ("OneHire", "Recruitment Portal",  "02 Feb 2026", "31 Jul 2026", "Not Started",
     "", "", "Yes"),
    ("Spot",    "City Module",         "02 Mar 2026", "30 Nov 2026", "Not Started",
     "", "", "Yes"),
    ("Pulse",   "Survey Analytics",    "16 Feb 2026", "31 Aug 2026", "Not Started",
     "", "", "Yes"),
]
for team, proj, sd, ed, phase, achieved, nxt, aligned in sample_projects:
    tc = TEAM_COLORS[team]
    ws3.cell(row=r, column=1, value=team)
    ws3.cell(row=r, column=1).font = Font(name="Aptos", bold=True, color=WHITE, size=10)
    ws3.cell(row=r, column=1).fill = PatternFill(start_color=tc, end_color=tc, fill_type="solid")
    ws3.cell(row=r, column=2, value=proj)
    ws3.cell(row=r, column=3, value=sd)
    ws3.cell(row=r, column=4, value=ed)
    ws3.cell(row=r, column=5, value=phase)
    ws3.cell(row=r, column=6, value=achieved)
    ws3.cell(row=r, column=7, value=nxt)
    ws3.cell(row=r, column=8, value=aligned)
    ws3.cell(row=r, column=9, value="")
    ws3.cell(row=r, column=10, value=TEAMS[team]["lead"])

    for c in range(1, max_c3 + 1):
        cell = ws3.cell(row=r, column=c)
        if c != 1:
            style_data_cell(cell, r, center=(c not in [2, 6, 7, 9]))
        else:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    if phase == "Post-MVP":
        for c in [5, 6]:
            cell = ws3.cell(row=r, column=c)
            cell.font = Font(name="Aptos", bold=True, size=10, color="065F46")
            cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

    ws3.row_dimensions[r].height = 48
    r += 1
    for c in range(1, max_c3 + 1):
        cell = ws3.cell(row=r, column=c)
        style_data_cell(cell, r)
    r += 1

ws3.freeze_panes = "A4"
auto_width(ws3)
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["F"].width = 48
ws3.column_dimensions["G"].width = 40
ws3.column_dimensions["I"].width = 28

# ───────────────────────────────────────────────────────────────────────
# SHEET 4 — SPRINT TRACKER (per team per sprint)
# ───────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Sprint Tracker")
ws4.sheet_properties.tabColor = GREEN

COLS4 = ["Team", "Sprint", "Dates", "Sprint Goal", "Achieved", "Tech Lead Feedback",
         "Achievement / Win", "Failure / Miss", "Blocker", "Status"]
max_c4 = len(COLS4)
add_title_block(ws4, "Sprint-by-Sprint Tracker — 2026", "Record goals, outcomes, tech lead feedback, and blockers every sprint", max_c4)
style_header_row(ws4, 3, max_c4, GREEN)
for i, h in enumerate(COLS4, 1):
    ws4.cell(row=3, column=i, value=h)

r = 4
for team in TEAMS:
    tc = TEAM_COLORS[team]
    for sp in SPRINTS_2026[:26]:
        sp_num = sp["num"]
        history = SPRINT_HISTORY.get((team, sp_num))
        current_sp = CURRENT_SPRINT_MAP.get(team, 1)
        is_completed = history is not None
        is_current = sp_num == current_sp

        ws4.cell(row=r, column=1, value=team)
        ws4.cell(row=r, column=1).font = Font(name="Aptos", bold=True, color=WHITE, size=10)
        ws4.cell(row=r, column=1).fill = PatternFill(start_color=tc, end_color=tc, fill_type="solid")
        ws4.cell(row=r, column=2, value=sp["name"])
        ws4.cell(row=r, column=3, value=f"{sp['start']} – {sp['end']}")

        if is_completed:
            ws4.cell(row=r, column=4, value=history["goal"])
            ws4.cell(row=r, column=5, value=history["achieved"])
            ws4.cell(row=r, column=6, value=history["feedback"])
            ws4.cell(row=r, column=7, value=history["win"])
            ws4.cell(row=r, column=8, value=history["failure"] or "—")
            ws4.cell(row=r, column=9, value=history["blocker"] or "—")
            ws4.cell(row=r, column=10, value="✅ Completed")
        elif is_current:
            ws4.cell(row=r, column=4, value="")
            ws4.cell(row=r, column=5, value="")
            ws4.cell(row=r, column=6, value="")
            ws4.cell(row=r, column=7, value="")
            ws4.cell(row=r, column=8, value="")
            ws4.cell(row=r, column=9, value="")
            ws4.cell(row=r, column=10, value="▶ In Progress")
        else:
            for c in range(4, max_c4 + 1):
                ws4.cell(row=r, column=c, value="")

        for c in range(1, max_c4 + 1):
            cell = ws4.cell(row=r, column=c)
            if c == 1:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_completed:
                style_completed_cell(cell, center=(c in [2, 3, 10]))
            elif is_current:
                cell.fill = PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid")
                cell.font = Font(name="Aptos", bold=True, size=10, color=DARK_GRAY)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if c in [2, 3, 10] else "left", vertical="center", wrap_text=True)
            else:
                style_data_cell(cell, r, center=(c in [2, 3, 10]))

        if is_completed and sp_num == 6:
            for c in [5, 7, 10]:
                cell = ws4.cell(row=r, column=c)
                cell.font = Font(name="Aptos", bold=True, size=10, color="065F46")
                cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

        r += 1

ws4.freeze_panes = "D4"
auto_width(ws4)
ws4.column_dimensions["D"].width = 36
ws4.column_dimensions["E"].width = 40
ws4.column_dimensions["F"].width = 34
ws4.column_dimensions["G"].width = 32
ws4.column_dimensions["H"].width = 22
ws4.column_dimensions["I"].width = 22

# ───────────────────────────────────────────────────────────────────────
# SHEET 5 — INDIVIDUAL MEMBER PERFORMANCE
# ───────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Member Performance")
ws5.sheet_properties.tabColor = AMBER

COLS5 = ["#", "Member", "Team(s)", "Role", "Current Sprint Task",
         "Progress %", "Tech Lead Rating", "Feedback / Notes", "Achievements (YTD)", "Areas to Improve"]
max_c5 = len(COLS5)
add_title_block(ws5, "Individual Member Performance — 2026", "Track each member's progress, ratings, and feedback across sprints", max_c5)
style_header_row(ws5, 3, max_c5, AMBER)
for i, h in enumerate(COLS5, 1):
    ws5.cell(row=3, column=i, value=h)

MVP_ACHIEVEMENT = "Contributed to successful MVP launch of 4 products (AG ONE, OneWork, Learn, Safe) in 6 sprints"

r = 4
for idx, member in enumerate(all_members, 1):
    teams_for = [t for t, info in TEAMS.items() if member in info["members"]]
    is_lead = any(TEAMS[t]["lead"] == member for t in teams_for)
    mvp_teams = [t for t in teams_for if t in ("AG ONE", "OneWork", "Learn", "Safe")]

    ws5.cell(row=r, column=1, value=idx)
    ws5.cell(row=r, column=2, value=member)
    ws5.cell(row=r, column=3, value=", ".join(teams_for))
    ws5.cell(row=r, column=4, value="Tech Lead" if is_lead else "Developer")
    ws5.cell(row=r, column=5, value="")
    ws5.cell(row=r, column=6, value="")
    ws5.cell(row=r, column=7, value="")
    ws5.cell(row=r, column=8, value="")
    ws5.cell(row=r, column=9, value=MVP_ACHIEVEMENT if mvp_teams else "")
    ws5.cell(row=r, column=10, value="")

    for c in range(1, max_c5 + 1):
        cell = ws5.cell(row=r, column=c)
        style_data_cell(cell, r, center=(c in [1, 4, 6, 7]))
        if c == 2:
            cell.font = Font(name="Aptos", bold=True, size=10, color=DARK_GRAY)
    if mvp_teams:
        cell9 = ws5.cell(row=r, column=9)
        cell9.font = Font(name="Aptos", bold=True, size=10, color="065F46")
        cell9.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    r += 1

ws5.freeze_panes = "C4"
auto_width(ws5)
ws5.column_dimensions["E"].width = 30
ws5.column_dimensions["H"].width = 32
ws5.column_dimensions["I"].width = 48
ws5.column_dimensions["J"].width = 28

# ───────────────────────────────────────────────────────────────────────
# SHEET 6 — YEARLY TARGETS & MILESTONES
# ───────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("2026 Targets")
ws6.sheet_properties.tabColor = RED

COLS6 = ["Team", "Q1 Target (Jan–Mar)", "Q1 Status", "Q2 Target (Apr–Jun)", "Q2 Status",
         "Q3 Target (Jul–Sep)", "Q3 Status", "Q4 Target (Oct–Dec)", "Q4 Status", "Year-End Goal"]
max_c6 = len(COLS6)
add_title_block(ws6, "2026 Yearly Targets & Quarterly Milestones", "High-level view of what each team must deliver per quarter", max_c6)
style_header_row(ws6, 3, max_c6, RED)
for i, h in enumerate(COLS6, 1):
    ws6.cell(row=3, column=i, value=h)

YEARLY_DATA = {
    "AG ONE": {
        "q1": "MVP launch — platform, users, roles, tenants, SSO, permissions",
        "q1s": "✅ Completed",
        "q2": "Platform hardening, advanced admin, analytics",
        "q2s": "Not Started",
        "q3": "Multi-product governance, API marketplace",
        "q3s": "Not Started",
        "q4": "Enterprise features, scale & performance",
        "q4s": "Not Started",
        "goal": "Full enterprise IAM platform with all products integrated",
    },
    "OneWork": {
        "q1": "MVP launch — employee mgmt, workflows, dashboards",
        "q1s": "✅ Completed",
        "q2": "Advanced workflows, integrations, reporting v2",
        "q2s": "Not Started",
        "q3": "Mobile support, AI-powered insights",
        "q3s": "Not Started",
        "q4": "Enterprise scale, advanced automation",
        "q4s": "Not Started",
        "goal": "Complete workforce management platform",
    },
    "Learn": {
        "q1": "MVP launch — courses, paths, assessments, certs",
        "q1s": "✅ Completed",
        "q2": "Advanced analytics, content marketplace",
        "q2s": "Not Started",
        "q3": "Mobile learning, AI recommendations",
        "q3s": "Not Started",
        "q4": "Enterprise LMS with full reporting",
        "q4s": "Not Started",
        "goal": "Full-featured LMS with marketplace & analytics",
    },
    "Safe": {
        "q1": "MVP launch — policies, compliance, audit, risk matrix",
        "q1s": "✅ Completed",
        "q2": "Regulatory automation, advanced reporting",
        "q2s": "Not Started",
        "q3": "Third-party integrations, automated alerts",
        "q3s": "Not Started",
        "q4": "Enterprise compliance suite",
        "q4s": "Not Started",
        "goal": "End-to-end compliance & risk management suite",
    },
    "OneHire": {
        "q1": "Requirements & design",
        "q1s": "Not Started",
        "q2": "MVP development — job postings, applicant tracking",
        "q2s": "Not Started",
        "q3": "MVP launch & post-launch improvements",
        "q3s": "Not Started",
        "q4": "Advanced hiring analytics & integrations",
        "q4s": "Not Started",
        "goal": "Full recruitment & applicant tracking system",
    },
    "Spot": {
        "q1": "Requirements & design",
        "q1s": "Not Started",
        "q2": "Core city module development",
        "q2s": "Not Started",
        "q3": "MVP launch — city management features",
        "q3s": "Not Started",
        "q4": "Advanced features & scale",
        "q4s": "Not Started",
        "goal": "City management module fully operational",
    },
    "Pulse": {
        "q1": "Requirements & design",
        "q1s": "Not Started",
        "q2": "Survey builder & distribution engine",
        "q2s": "Not Started",
        "q3": "MVP launch — survey analytics & dashboards",
        "q3s": "Not Started",
        "q4": "AI-powered insights & benchmarking",
        "q4s": "Not Started",
        "goal": "Complete employee engagement & survey platform",
    },
}

r = 4
for team in TEAMS:
    tc = TEAM_COLORS[team]
    yd = YEARLY_DATA[team]
    ws6.cell(row=r, column=1, value=team)
    ws6.cell(row=r, column=1).font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    ws6.cell(row=r, column=1).fill = PatternFill(start_color=tc, end_color=tc, fill_type="solid")
    ws6.cell(row=r, column=2, value=yd["q1"])
    ws6.cell(row=r, column=3, value=yd["q1s"])
    ws6.cell(row=r, column=4, value=yd["q2"])
    ws6.cell(row=r, column=5, value=yd["q2s"])
    ws6.cell(row=r, column=6, value=yd["q3"])
    ws6.cell(row=r, column=7, value=yd["q3s"])
    ws6.cell(row=r, column=8, value=yd["q4"])
    ws6.cell(row=r, column=9, value=yd["q4s"])
    ws6.cell(row=r, column=10, value=yd["goal"])

    for c in range(1, max_c6 + 1):
        cell = ws6.cell(row=r, column=c)
        if c == 1:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if c in [3,5,7,9] else "left", vertical="center", wrap_text=True)
            cell.font = Font(name="Aptos", size=10, color=DARK_GRAY)

    for sc in [3, 5, 7, 9]:
        cell = ws6.cell(row=r, column=sc)
        val = cell.value
        if val == "✅ Completed":
            cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
            cell.font = Font(name="Aptos", bold=True, size=10, color="065F46")
        else:
            cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
            cell.font = Font(name="Aptos", size=10, italic=True, color=MED_GRAY)

    ws6.row_dimensions[r].height = 56
    r += 1

ws6.freeze_panes = "B4"
auto_width(ws6)
for col_letter in ["B", "D", "F", "H", "J"]:
    ws6.column_dimensions[col_letter].width = 36
for col_letter in ["C", "E", "G", "I"]:
    ws6.column_dimensions[col_letter].width = 16

# ───────────────────────────────────────────────────────────────────────
# SHEET 7 — EXTERNAL / OUTSOURCE PROJECTS
# ───────────────────────────────────────────────────────────────────────
ws_ext = wb.create_sheet("External Projects")
ws_ext.sheet_properties.tabColor = EXTERNAL_COLOR

COLS_EXT = ["#", "Project Name", "Client / Partner", "Assigned Resources",
            "Role / Responsibility", "Allocation %", "Start Date", "End Date",
            "Current Phase", "Status", "Key Deliverables", "Notes / Risk"]
max_ext = len(COLS_EXT)
add_title_block(ws_ext, "External / Outsource Projects — 2026",
                "Track resources assigned to external projects, outsourced work, and partner engagements", max_ext)
style_header_row(ws_ext, 3, max_ext, EXTERNAL_COLOR)
for i, h in enumerate(COLS_EXT, 1):
    ws_ext.cell(row=3, column=i, value=h)

r = 4
for eidx, (ep_name, ep) in enumerate(EXTERNAL_PROJECTS.items(), 1):
    ws_ext.cell(row=r, column=1, value=eidx)
    ws_ext.cell(row=r, column=2, value=ep_name)
    ws_ext.cell(row=r, column=2).font = Font(name="Aptos", bold=True, size=11, color=WHITE)
    ws_ext.cell(row=r, column=2).fill = PatternFill(start_color=ep["color"], end_color=ep["color"], fill_type="solid")
    ws_ext.cell(row=r, column=3, value=ep.get("client", ""))
    ws_ext.cell(row=r, column=4, value=", ".join(ep["members"]) if ep["members"] else "")
    ws_ext.cell(row=r, column=5, value="")
    ws_ext.cell(row=r, column=6, value="")
    ws_ext.cell(row=r, column=7, value=ep.get("start", ""))
    ws_ext.cell(row=r, column=8, value=ep.get("end", ""))
    ws_ext.cell(row=r, column=9, value="")
    ws_ext.cell(row=r, column=10, value=ep.get("status", ""))
    ws_ext.cell(row=r, column=11, value=ep.get("description", ""))
    ws_ext.cell(row=r, column=12, value=ep.get("notes", ""))

    for c in range(1, max_ext + 1):
        cell = ws_ext.cell(row=r, column=c)
        if c != 2:
            style_data_cell(cell, r, center=(c in [1, 6, 7, 8, 10]))
        else:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    if ep.get("status") == "Active":
        ws_ext.cell(row=r, column=10).fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
        ws_ext.cell(row=r, column=10).font = Font(name="Aptos", bold=True, size=10, color="065F46")
    r += 1

# Pre-fill 9 more blank rows for future external projects
for extra in range(9):
    ws_ext.cell(row=r, column=1, value=eidx + extra + 1)
    for c in range(1, max_ext + 1):
        cell = ws_ext.cell(row=r, column=c)
        style_data_cell(cell, r, center=(c in [1, 6, 7, 8, 10]))
    r += 1

ws_ext.freeze_panes = "C4"
auto_width(ws_ext)
ws_ext.column_dimensions["B"].width = 22
ws_ext.column_dimensions["C"].width = 20
ws_ext.column_dimensions["D"].width = 30
ws_ext.column_dimensions["E"].width = 24
ws_ext.column_dimensions["K"].width = 34
ws_ext.column_dimensions["L"].width = 30

# ═══════════════════════════════════════════════════════════════════════
# RELEASE TRACKING DATA — all products, full history
# ═══════════════════════════════════════════════════════════════════════

RELEASE_TYPE_COLORS = {
    "MVP":       ("10B981", "D1FAE5", "065F46"),
    "Major":     ("1E3A5F", "DBEAFE", "1E3A5F"),
    "Minor":     ("3B82F6", "DBEAFE", "1E3A5F"),
    "Hotfix":    ("EF4444", "FEE2E2", "991B1B"),
    "Patch":     ("F59E0B", "FEF3C7", "92400E"),
    "Sprint":    ("8B5CF6", "EDE9FE", "5B21B6"),
    "Planned":   ("6B7280", "F3F4F6", "374151"),
}

# ── RELEASE HISTORY — every release that has happened or is planned ───
RELEASE_HISTORY = [
    # AG ONE
    {"product": "AG ONE",  "version": "v1.0.0", "name": "AG ONE MVP",                "type": "MVP",    "date": "29 Mar 2026", "sprint": "Sprint 6", "status": "✅ Released", "highlights": "Full IAM platform: Users, Roles, Tenants, SSO, Permissions, Audit, Assign Access"},
    {"product": "AG ONE",  "version": "v1.0.1", "name": "AG ONE Hotfix 1",           "type": "Hotfix", "date": "02 Apr 2026", "sprint": "Sprint 7", "status": "✅ Released", "highlights": "SSO token refresh loop fix, role sync cache invalidation, tenant switching fix"},
    {"product": "AG ONE",  "version": "v1.0.2", "name": "AG ONE Hotfix 2",           "type": "Hotfix", "date": "08 Apr 2026", "sprint": "Sprint 7", "status": "✅ Released", "highlights": "Audit log pagination, API key duplicate name validation, timezone fix"},
    {"product": "AG ONE",  "version": "v1.1.0", "name": "AG ONE Q2 Enhancement",     "type": "Minor",  "date": "May 2026",    "sprint": "Sprint 9–10", "status": "📋 Planned", "highlights": "Advanced user analytics, bulk role operations, enhanced audit dashboard"},
    {"product": "AG ONE",  "version": "v1.2.0", "name": "AG ONE Multi-Product Admin", "type": "Minor", "date": "Jul 2026",    "sprint": "Sprint 13–14", "status": "📋 Planned", "highlights": "Multi-product admin console, cross-product role mapping, API marketplace v1"},
    {"product": "AG ONE",  "version": "v2.0.0", "name": "AG ONE Enterprise",         "type": "Major",  "date": "Oct 2026",    "sprint": "Sprint 20–22", "status": "📋 Planned", "highlights": "Enterprise SSO (SAML+OIDC), advanced RBAC, white-label branding, scale"},

    # OneWork
    {"product": "OneWork", "version": "v1.0.0", "name": "OneWork MVP",               "type": "MVP",    "date": "29 Mar 2026", "sprint": "Sprint 6", "status": "✅ Released", "highlights": "Employee mgmt, workflows, task mgmt, dashboards, reporting"},
    {"product": "OneWork", "version": "v1.1.0", "name": "OneWork Q2 Enhancement",    "type": "Minor",  "date": "Jun 2026",    "sprint": "Sprint 11–12", "status": "📋 Planned", "highlights": "Advanced workflow templates, email notifications, calendar integration"},
    {"product": "OneWork", "version": "v1.2.0", "name": "OneWork Reporting v2",      "type": "Minor",  "date": "Aug 2026",    "sprint": "Sprint 15–16", "status": "📋 Planned", "highlights": "Reporting v2 with charts, mobile-responsive, third-party integrations"},
    {"product": "OneWork", "version": "v2.0.0", "name": "OneWork Enterprise",        "type": "Major",  "date": "Nov 2026",    "sprint": "Sprint 21–23", "status": "📋 Planned", "highlights": "AI-powered insights, advanced automation, enterprise scale"},

    # Learn
    {"product": "Learn",   "version": "v1.0.0", "name": "Learn MVP",                 "type": "MVP",    "date": "29 Mar 2026", "sprint": "Sprint 6", "status": "✅ Released", "highlights": "Courses, learning paths, assessments, certifications, progress tracking"},
    {"product": "Learn",   "version": "v1.1.0", "name": "Learn Analytics",           "type": "Minor",  "date": "Jun 2026",    "sprint": "Sprint 11–12", "status": "📋 Planned", "highlights": "Advanced analytics dashboard, learner recommendations, completion reports"},
    {"product": "Learn",   "version": "v1.2.0", "name": "Learn Marketplace",         "type": "Minor",  "date": "Aug 2026",    "sprint": "Sprint 15–16", "status": "📋 Planned", "highlights": "Content marketplace, SCORM support, mobile-responsive learning"},
    {"product": "Learn",   "version": "v2.0.0", "name": "Learn Enterprise",          "type": "Major",  "date": "Nov 2026",    "sprint": "Sprint 21–23", "status": "📋 Planned", "highlights": "AI recommendations, adaptive learning, full mobile app"},

    # Safe
    {"product": "Safe",    "version": "v1.0.0", "name": "Safe MVP",                   "type": "MVP",    "date": "29 Mar 2026", "sprint": "Sprint 6", "status": "✅ Released", "highlights": "Policies, compliance tracking, audit workflows, risk matrix, dashboards"},
    {"product": "Safe",    "version": "v1.1.0", "name": "Safe Regulatory Auto",      "type": "Minor",  "date": "Jun 2026",    "sprint": "Sprint 11–12", "status": "📋 Planned", "highlights": "Regulatory update automation, scheduled compliance checks, advanced reporting"},
    {"product": "Safe",    "version": "v1.2.0", "name": "Safe Integrations",         "type": "Minor",  "date": "Sep 2026",    "sprint": "Sprint 17–18", "status": "📋 Planned", "highlights": "Third-party GRC integrations, automated alerts, evidence collection"},
    {"product": "Safe",    "version": "v2.0.0", "name": "Safe Enterprise",           "type": "Major",  "date": "Dec 2026",    "sprint": "Sprint 24–26", "status": "📋 Planned", "highlights": "Enterprise compliance suite with AI-driven risk predictions"},

    # OneHire
    {"product": "OneHire", "version": "v0.1.0", "name": "OneHire Alpha",             "type": "Sprint", "date": "May 2026",    "sprint": "Sprint 9–10", "status": "📋 Planned", "highlights": "Core job posting and applicant tracking features"},
    {"product": "OneHire", "version": "v0.5.0", "name": "OneHire Beta",              "type": "Sprint", "date": "Jun 2026",    "sprint": "Sprint 11–12", "status": "📋 Planned", "highlights": "Interview scheduling, candidate pipeline, basic reporting"},
    {"product": "OneHire", "version": "v1.0.0", "name": "OneHire MVP",               "type": "MVP",    "date": "Jul 2026",    "sprint": "Sprint 13–14", "status": "📋 Planned", "highlights": "Full recruitment portal with applicant tracking and analytics"},

    # Spot
    {"product": "Spot",    "version": "v0.1.0", "name": "Spot Alpha",                "type": "Sprint", "date": "May 2026",    "sprint": "Sprint 9–10", "status": "📋 Planned", "highlights": "Core city module, location management"},
    {"product": "Spot",    "version": "v1.0.0", "name": "Spot MVP",                  "type": "MVP",    "date": "Sep 2026",    "sprint": "Sprint 17–18", "status": "📋 Planned", "highlights": "City management, geolocation features, admin dashboard"},

    # Pulse
    {"product": "Pulse",   "version": "v0.1.0", "name": "Pulse Alpha",               "type": "Sprint", "date": "May 2026",    "sprint": "Sprint 9–10", "status": "📋 Planned", "highlights": "Survey builder, basic distribution"},
    {"product": "Pulse",   "version": "v1.0.0", "name": "Pulse MVP",                 "type": "MVP",    "date": "Sep 2026",    "sprint": "Sprint 17–18", "status": "📋 Planned", "highlights": "Survey analytics, employee engagement dashboards, benchmarking"},
]

# ───────────────────────────────────────────────────────────────────────
# SHEET 8 — RELEASE DASHBOARD (overview summary)
# ───────────────────────────────────────────────────────────────────────
ws_rd = wb.create_sheet("Release Dashboard")
ws_rd.sheet_properties.tabColor = "0EA5E9"

COLS_RD = ["Product", "Latest Version", "Latest Release", "Release Date",
           "Total Releases", "Hotfixes", "Next Planned", "Next Version", "Target Date"]
max_rd = len(COLS_RD)
add_title_block(ws_rd, "Release Dashboard — All Products 2026",
                "At-a-glance: latest release, total count, and next planned for every product", max_rd)
style_header_row(ws_rd, 3, max_rd, "0EA5E9")
for i, h in enumerate(COLS_RD, 1):
    ws_rd.cell(row=3, column=i, value=h)

# Versioning standard banner
r = 4
ws_rd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_rd)
banner = ws_rd.cell(row=r, column=1,
    value="📏  Versioning Standard: vMAJOR.MINOR.PATCH  •  Types: MVP | Major | Minor | Hotfix | Patch | Sprint  •  Quarterly releases + hotfixes as needed")
banner.font = Font(name="Aptos", bold=True, size=10, color="0369A1")
banner.fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
banner.alignment = Alignment(horizontal="center", vertical="center")
banner.border = thin_border
for c in range(2, max_rd + 1):
    ws_rd.cell(row=r, column=c).fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    ws_rd.cell(row=r, column=c).border = thin_border
ws_rd.row_dimensions[r].height = 28
r += 1

for product in TEAMS.keys():
    tc = TEAM_COLORS[product]
    prod_releases = [x for x in RELEASE_HISTORY if x["product"] == product]
    released = [x for x in prod_releases if "Released" in x["status"]]
    hotfixes = [x for x in released if x["type"] == "Hotfix"]
    planned = [x for x in prod_releases if "Planned" in x["status"]]

    latest = released[-1] if released else None
    nxt = planned[0] if planned else None

    ws_rd.cell(row=r, column=1, value=product)
    ws_rd.cell(row=r, column=1).font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    ws_rd.cell(row=r, column=1).fill = PatternFill(start_color=tc, end_color=tc, fill_type="solid")
    ws_rd.cell(row=r, column=2, value=latest["version"] if latest else "—")
    ws_rd.cell(row=r, column=3, value=latest["name"] if latest else "—")
    ws_rd.cell(row=r, column=4, value=latest["date"] if latest else "—")
    ws_rd.cell(row=r, column=5, value=len(released))
    ws_rd.cell(row=r, column=6, value=len(hotfixes))
    ws_rd.cell(row=r, column=7, value=nxt["name"] if nxt else "—")
    ws_rd.cell(row=r, column=8, value=nxt["version"] if nxt else "—")
    ws_rd.cell(row=r, column=9, value=nxt["date"] if nxt else "—")

    for c in range(1, max_rd + 1):
        cell = ws_rd.cell(row=r, column=c)
        if c != 1:
            style_data_cell(cell, r, center=True)
        else:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Highlight released count
    if len(released) > 0:
        ws_rd.cell(row=r, column=5).font = Font(name="Aptos", bold=True, size=11, color="065F46")
        ws_rd.cell(row=r, column=5).fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    if len(hotfixes) > 0:
        ws_rd.cell(row=r, column=6).font = Font(name="Aptos", bold=True, size=10, color="991B1B")
        ws_rd.cell(row=r, column=6).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

    ws_rd.row_dimensions[r].height = 32
    r += 1

# Summary row
r += 1
ws_rd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws_rd.cell(row=r, column=1, value="TOTAL ACROSS ALL PRODUCTS")
ws_rd.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=11, color=WHITE)
ws_rd.cell(row=r, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
for c in range(2, 5):
    ws_rd.cell(row=r, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
total_released = len([x for x in RELEASE_HISTORY if "Released" in x["status"]])
total_hotfixes = len([x for x in RELEASE_HISTORY if "Released" in x["status"] and x["type"] == "Hotfix"])
total_planned = len([x for x in RELEASE_HISTORY if "Planned" in x["status"]])
ws_rd.cell(row=r, column=5, value=total_released)
ws_rd.cell(row=r, column=5).font = Font(name="Aptos", bold=True, size=12, color="065F46")
ws_rd.cell(row=r, column=5).fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
ws_rd.cell(row=r, column=6, value=total_hotfixes)
ws_rd.cell(row=r, column=6).font = Font(name="Aptos", bold=True, size=12, color="991B1B")
ws_rd.cell(row=r, column=6).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
ws_rd.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
ws_rd.cell(row=r, column=7, value=f"{total_planned} releases planned")
ws_rd.cell(row=r, column=7).font = Font(name="Aptos", bold=True, size=11, color=MED_GRAY)
for c in range(5, max_rd + 1):
    ws_rd.cell(row=r, column=c).border = thin_border
    ws_rd.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

ws_rd.freeze_panes = "B5"
auto_width(ws_rd)
ws_rd.column_dimensions["C"].width = 28
ws_rd.column_dimensions["G"].width = 28
ws_rd.column_dimensions["H"].width = 14

# ───────────────────────────────────────────────────────────────────────
# SHEET 9 — RELEASE HISTORY (all products, grouped by product)
# ───────────────────────────────────────────────────────────────────────
ws_rh = wb.create_sheet("Release History")
ws_rh.sheet_properties.tabColor = "0284C7"

COLS_RH = ["#", "Product", "Version", "Release Name", "Type", "Date",
           "Sprint", "Status", "Highlights / What's Included"]
max_rh = len(COLS_RH)
add_title_block(ws_rh, "Release History — All Products 2026",
                "Every release grouped by product: Released ✅ + Upcoming 📋  •  Scroll down for each team", max_rh)

r = 3
global_idx = 0

for product in TEAMS.keys():
    tc = TEAM_COLORS[product]
    tl = TEAM_LIGHT[product]
    prod_releases = [x for x in RELEASE_HISTORY if x["product"] == product]
    released = [x for x in prod_releases if "Released" in x["status"]]
    planned = [x for x in prod_releases if "Planned" in x["status"]]
    hotfixes = [x for x in released if x["type"] == "Hotfix"]

    # Product group header
    ws_rh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_rh)
    stats = (f"{product}  —  Released: {len(released)}  •  Hotfixes: {len(hotfixes)}  •  "
             f"Planned: {len(planned)}  •  Latest: {released[-1]['version'] if released else '—'}  •  "
             f"Next: {planned[0]['version'] if planned else '—'}")
    ws_rh.cell(row=r, column=1, value=stats)
    ws_rh.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=11, color=WHITE)
    ws_rh.cell(row=r, column=1).fill = PatternFill(start_color=tc, end_color=tc, fill_type="solid")
    ws_rh.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws_rh.cell(row=r, column=1).border = thin_border
    for c in range(2, max_rh + 1):
        ws_rh.cell(row=r, column=c).fill = PatternFill(start_color=tc, end_color=tc, fill_type="solid")
        ws_rh.cell(row=r, column=c).border = thin_border
    ws_rh.row_dimensions[r].height = 30
    r += 1

    # Column headers for this group
    style_header_row(ws_rh, r, max_rh, "475569")
    for i, h in enumerate(COLS_RH, 1):
        ws_rh.cell(row=r, column=i, value=h)
    r += 1

    # All releases for this product (released first, then planned)
    for rel in prod_releases:
        global_idx += 1
        rt_colors = RELEASE_TYPE_COLORS.get(rel["type"], ("6B7280", "F3F4F6", "374151"))
        is_released = "Released" in rel["status"]

        ws_rh.cell(row=r, column=1, value=global_idx)
        ws_rh.cell(row=r, column=2, value=rel["product"])
        ws_rh.cell(row=r, column=3, value=rel["version"])
        ws_rh.cell(row=r, column=3).font = Font(name="Aptos", bold=True, size=10, color=DARK_GRAY)
        ws_rh.cell(row=r, column=4, value=rel["name"])
        ws_rh.cell(row=r, column=5, value=rel["type"])
        ws_rh.cell(row=r, column=5).font = Font(name="Aptos", bold=True, size=10, color=rt_colors[2])
        ws_rh.cell(row=r, column=5).fill = PatternFill(start_color=rt_colors[1], end_color=rt_colors[1], fill_type="solid")
        ws_rh.cell(row=r, column=6, value=rel["date"])
        ws_rh.cell(row=r, column=7, value=rel["sprint"])
        ws_rh.cell(row=r, column=8, value=rel["status"])
        ws_rh.cell(row=r, column=9, value=rel["highlights"])

        for c in range(1, max_rh + 1):
            cell = ws_rh.cell(row=r, column=c)
            if c not in [5]:
                style_data_cell(cell, r, center=(c not in [4, 9]))
            else:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Status styling
        status_cell = ws_rh.cell(row=r, column=8)
        if is_released:
            status_cell.font = Font(name="Aptos", bold=True, size=10, color="065F46")
            status_cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
        else:
            status_cell.font = Font(name="Aptos", size=10, italic=True, color=MED_GRAY)
            status_cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

        ws_rh.row_dimensions[r].height = 30
        r += 1

    # 3 blank rows for future entries per product
    for _ in range(3):
        global_idx += 1
        ws_rh.cell(row=r, column=1, value="")
        ws_rh.cell(row=r, column=2, value=product)
        for c in range(1, max_rh + 1):
            cell = ws_rh.cell(row=r, column=c)
            style_data_cell(cell, r, center=(c not in [4, 9]))
        ws_rh.cell(row=r, column=2).font = Font(name="Aptos", size=10, italic=True, color=MED_GRAY)
        r += 1

    r += 1  # gap between products

ws_rh.freeze_panes = "C4"
auto_width(ws_rh)
ws_rh.column_dimensions["D"].width = 30
ws_rh.column_dimensions["I"].width = 55

# ───────────────────────────────────────────────────────────────────────
# SHEET 10 — RELEASE NAMING STANDARD & TYPES
# ───────────────────────────────────────────────────────────────────────
ws_rs = wb.create_sheet("Release Standard")
ws_rs.sheet_properties.tabColor = "0369A1"

max_rs = 4
add_title_block(ws_rs, "Release Naming Standard & Versioning Policy",
                "Share this with all teams so everyone follows the same release naming and numbering", max_rs)

# Versioning format
r = 3
style_header_row(ws_rs, r, max_rs, NAVY)
ws_rs.cell(row=r, column=1, value="Versioning Format")
ws_rs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_rs)

r = 4
version_format = [
    ("Format", "vMAJOR.MINOR.PATCH", "Example: v1.2.3"),
    ("MAJOR", "Breaking changes or major new feature sets", "v1.0.0 → v2.0.0"),
    ("MINOR", "New features, enhancements (backward compatible)", "v1.0.0 → v1.1.0"),
    ("PATCH", "Bug fixes, hotfixes, minor corrections", "v1.0.0 → v1.0.1"),
]
for label, desc, example in version_format:
    ws_rs.cell(row=r, column=1, value=label)
    ws_rs.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=11, color=WHITE)
    ws_rs.cell(row=r, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws_rs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws_rs.cell(row=r, column=2, value=desc)
    ws_rs.cell(row=r, column=2).font = Font(name="Aptos", size=10, color=DARK_GRAY)
    ws_rs.cell(row=r, column=4, value=example)
    ws_rs.cell(row=r, column=4).font = Font(name="Aptos", size=10, italic=True, color=MED_GRAY)
    for c in range(1, max_rs + 1):
        ws_rs.cell(row=r, column=c).border = thin_border
        ws_rs.cell(row=r, column=c).alignment = Alignment(horizontal="left" if c > 1 else "center", vertical="center")
    r += 1

r += 1
style_header_row(ws_rs, r, max_rs, "0EA5E9")
ws_rs.cell(row=r, column=1, value="Release Type")
ws_rs.cell(row=r, column=2, value="Description")
ws_rs.cell(row=r, column=3, value="Frequency")
ws_rs.cell(row=r, column=4, value="Approval")
r += 1

release_types = [
    ("MVP", "First production release of a product", "Once per product", "Director + Tech Lead", "10B981"),
    ("Major", "Significant new features or breaking changes", "1–2 per year", "Director + Tech Lead", "1E3A5F"),
    ("Minor", "New features & enhancements, backward compatible", "Quarterly", "Tech Lead", "3B82F6"),
    ("Hotfix", "Critical production fix, expedited deployment", "As needed (urgent)", "Tech Lead + post-mortem", "EF4444"),
    ("Patch", "Non-critical bug fixes, scheduled deployment", "As needed", "Tech Lead", "F59E0B"),
    ("Sprint", "Sprint-based release (pre-MVP iterations)", "Every 2 weeks", "Tech Lead", "8B5CF6"),
]

for rtype, desc, freq, approval, color in release_types:
    ws_rs.cell(row=r, column=1, value=rtype)
    ws_rs.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=10, color=WHITE)
    ws_rs.cell(row=r, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws_rs.cell(row=r, column=2, value=desc)
    ws_rs.cell(row=r, column=3, value=freq)
    ws_rs.cell(row=r, column=4, value=approval)
    for c in range(1, max_rs + 1):
        cell = ws_rs.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if c == 1 else "left", vertical="center", wrap_text=True)
        if c > 1:
            cell.font = Font(name="Aptos", size=10, color=DARK_GRAY)
    r += 1

r += 1
style_header_row(ws_rs, r, max_rs, NAVY)
ws_rs.cell(row=r, column=1, value="Release Naming Convention")
ws_rs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_rs)
r += 1

naming_examples = [
    ("Pattern", "[Product] [Type] — [Short Description]", "Used in Release Name column"),
    ("MVP", "AG ONE MVP", "First production release"),
    ("Minor", "AG ONE Q2 Enhancement", "Quarterly feature release"),
    ("Hotfix", "AG ONE Hotfix 1", "Numbered sequentially"),
    ("Major", "AG ONE Enterprise", "Named for the theme"),
    ("Sprint", "OneHire Alpha / OneHire Beta", "Pre-MVP iterations"),
]

for label, example, note in naming_examples:
    ws_rs.cell(row=r, column=1, value=label)
    ws_rs.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=10, color=NAVY)
    ws_rs.cell(row=r, column=2, value=example)
    ws_rs.cell(row=r, column=2).font = Font(name="Aptos", bold=True, size=10, color=DARK_GRAY)
    ws_rs.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws_rs.cell(row=r, column=3, value=note)
    ws_rs.cell(row=r, column=3).font = Font(name="Aptos", size=10, italic=True, color=MED_GRAY)
    for c in range(1, max_rs + 1):
        ws_rs.cell(row=r, column=c).border = thin_border
        ws_rs.cell(row=r, column=c).alignment = Alignment(vertical="center")
    r += 1

ws_rs.freeze_panes = "A4"
auto_width(ws_rs)
ws_rs.column_dimensions["A"].width = 16
ws_rs.column_dimensions["B"].width = 44
ws_rs.column_dimensions["C"].width = 26
ws_rs.column_dimensions["D"].width = 28

# ───────────────────────────────────────────────────────────────────────
# SHEET 11 — LEGEND & INSTRUCTIONS
# ───────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("How To Use")
ws7.sheet_properties.tabColor = MED_GRAY

max_c7 = 4
add_title_block(ws7, "How To Use This Tracker", "Quick guide for all stakeholders", max_c7)

instructions = [
    ("Sheet", "Purpose", "Update Frequency", "Who Updates"),
    ("Team Overview", "See all 7 internal teams, current sprint, status at a glance", "Every sprint", "Engineering Manager"),
    ("Resource Allocation", "Internal + external assignments per member, shared & outsource flags", "When team changes", "Engineering Manager"),
    ("Project Roadmap", "Track project start/end, achievements, next targets, alignment", "Every sprint", "Tech Leads"),
    ("Sprint Tracker", "Detailed sprint goals, outcomes, tech lead feedback, blockers", "Every sprint", "Tech Leads"),
    ("Member Performance", "Individual progress, ratings, achievements", "Every sprint", "Tech Leads"),
    ("2026 Targets", "Quarterly milestones and year-end goals per team", "Quarterly", "Engineering Manager"),
    ("External Projects", "Track resources on external / outsource / partner projects", "When assignments change", "Engineering Manager"),
    ("Release Dashboard", "At-a-glance: latest release, totals, and next planned per product", "Every release", "Engineering Manager"),
    ("Release History", "All releases grouped by product — released + upcoming with highlights", "Every release", "Tech Leads"),
    ("Release Standard", "Versioning policy, naming convention, release types — share with all teams", "Reference only", "—"),
]

style_header_row(ws7, 3, max_c7, NAVY)
for i, h in enumerate(instructions[0], 1):
    ws7.cell(row=3, column=i, value=h)

for ri, row_data in enumerate(instructions[1:], 4):
    for ci, val in enumerate(row_data, 1):
        cell = ws7.cell(row=ri, column=ci, value=val)
        style_data_cell(cell, ri, center=(ci != 2))

r_leg = len(instructions) + 4
ws7.cell(row=r_leg, column=1, value="STATUS LEGEND")
ws7.cell(row=r_leg, column=1).font = Font(name="Aptos", bold=True, size=12, color=NAVY)
legend = [
    ("✅ Completed", NAVY, LIGHT_BLUE),
    ("▶ In Progress / On Track", GREEN, LIGHT_GREEN),
    ("At Risk", AMBER, LIGHT_AMBER),
    ("Blocked", RED, LIGHT_RED),
    ("Not Started", MED_GRAY, LIGHT_GRAY),
]
for li, (label, fg, bg) in enumerate(legend, r_leg + 1):
    cell = ws7.cell(row=li, column=1, value=f"  ●  {label}")
    cell.font = Font(name="Aptos", bold=True, size=10, color=fg)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.border = thin_border
    ws7.merge_cells(start_row=li, start_column=1, end_row=li, end_column=2)

auto_width(ws7)
ws7.column_dimensions["B"].width = 52

# ─── PRINT SETTINGS & FINAL TOUCHES ──────────────────────────────────
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

# ─── SAVE ─────────────────────────────────────────────────────────────
output_path = "/workspace/AG_ONE_Team_Tracker_2026.xlsx"
wb.save(output_path)
print(f"✅ Saved: {output_path}")
