#!/usr/bin/env python3
"""Generate a full-year monthly tech team sync Excel workbook."""

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

YEAR = 2027
MONTHS = [
    ("January", "Jan"),
    ("February", "Feb"),
    ("March", "Mar"),
    ("April", "Apr"),
    ("May", "May"),
    ("June", "Jun"),
    ("July", "Jul"),
    ("August", "Aug"),
    ("September", "Sep"),
    ("October", "Oct"),
    ("November", "Nov"),
    ("December", "Dec"),
]

# Placeholder roster — replace with your actual team
DEFAULT_TEAM = [
    ("Member 1", "Backend Developer", "Core Platform"),
    ("Member 2", "Frontend Developer", "Web Apps"),
    ("Member 3", "Full Stack Developer", "Integrations"),
    ("Member 4", "DevOps Engineer", "Infrastructure"),
    ("Member 5", "QA Engineer", "Quality"),
    ("Member 6", "Tech Lead", "Architecture"),
]

NAVY = "1F4E79"
BLUE = "2E75B6"
TEAL = "1B7A6E"
LIGHT_BLUE = "D9E2F3"
LIGHT_TEAL = "D5F0ED"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"
ORANGE = "FCE4D6"
ALT = "F2F2F2"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_TEAL)
ALT_FILL = PatternFill("solid", fgColor=ALT)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_cell(cell, *, bold=False, fill=None, align="left", size=11, color="000000", wrap=True):
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = fill


def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_list_validation(ws, cell_range, options):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = f"Choose: {', '.join(options)}"
    dv.errorTitle = "Invalid selection"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_status_colors(ws, cell_range):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"In Progress"'], fill=PatternFill("solid", fgColor=YELLOW)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Not Started"'], fill=PatternFill("solid", fgColor=ORANGE)),
    )


def build_instructions_sheet(ws):
    ws.title = "Start Here"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"TECH TEAM MONTHLY SYNC — {YEAR} WORKBOOK"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=18, color=WHITE)

    sections = [
        ("Purpose", [
            "Run consistent monthly sync sessions with every tech team member.",
            "Capture updates, feedback, gaps, and next-month priorities in one place.",
            "Share this workbook with the team before each sync so they can pre-fill their sections.",
        ]),
        ("How to use (Manager)", [
            "1. Update the Team Roster sheet with names, roles, and squads.",
            "2. Before each monthly sync, share the relevant month tab with the team.",
            "3. Ask each member to fill: Updates, New Items, What's Next, Gaps, and Feedback Needed.",
            "4. During the sync, review the Year Dashboard and Action Items Tracker.",
            "5. After the sync, log decisions, update statuses, and draft next month's agenda.",
        ]),
        ("How to use (Team Members)", [
            "1. Open your month's tab (e.g. 'Mar 2027').",
            "2. Find your row in the Member Updates section.",
            "3. Fill in all columns honestly — this drives the 1:1 and team discussion.",
            "4. Add any feedback or support requests in the Feedback Needed column.",
            "5. Submit at least 2 business days before the monthly sync meeting.",
        ]),
        ("Sheet guide", [
            "Year Dashboard — 12-month calendar, sync dates, themes, and status at a glance.",
            "Team Roster — Master list of all tech members (edit names here first).",
            "Jan–Dec tabs — Monthly sync workspace per member.",
            "Action Items Tracker — Cross-month follow-ups with owners and due dates.",
            "Feedback Log — Team feedback, concerns, and manager responses.",
            "Gap & Growth Tracker — Skills, process, and resource gaps to address.",
            "Agenda Library — Reusable agenda items for monthly syncs.",
        ]),
        ("Monthly sync agenda (recommended 60 min)", [
            "0–5 min   | Opening — goals for this month, review last month's action items.",
            "5–25 min  | Round-robin member updates (2–3 min each): highlights, blockers, what's next.",
            "25–40 min | Deep dive on top 2–3 team gaps or cross-team dependencies.",
            "40–50 min | Feedback round — what to start, stop, continue; process improvements.",
            "50–60 min | Close — confirm action items, owners, due dates; preview next month agenda.",
        ]),
        ("Status legend", [
            "Not Started — Not yet begun.",
            "In Progress — Actively being worked on.",
            "Blocked — Cannot proceed without help or a decision.",
            "Done — Completed this month.",
        ]),
    ]

    row = 3
    for title, lines in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, title.upper())
        style_cell(ws.cell(row, 1), bold=True, fill=SUBHEADER_FILL, size=12)
        row += 1
        for line in lines:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row, 1, line)
            style_cell(ws.cell(row, 1))
            row += 1
        row += 1

    set_col_widths(ws, [18, 18, 18, 18, 18, 18])
    ws.sheet_view.zoomScale = 100


def build_roster_sheet(ws):
    ws.title = "Team Roster"
    ws.merge_cells("A1:G1")
    ws["A1"] = "TECH TEAM ROSTER — Edit names, roles, and squads here"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    headers = ["#", "Name", "Role", "Squad / Area", "Email", "Active?", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=SUBHEADER_FILL, align="center")

    for i, (name, role, squad) in enumerate(DEFAULT_TEAM, 1):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [i, name, role, squad, "", "Yes", ""]
        for col, val in enumerate(values, 1):
            ws.cell(row, col, val)
            style_cell(ws.cell(row, col), fill=fill, align="center" if col in (1, 6) else "left")

    # Extra blank rows for more members
    for i in range(len(DEFAULT_TEAM) + 1, len(DEFAULT_TEAM) + 9):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 1, i)
        for col in range(1, 8):
            style_cell(ws.cell(row, col), fill=fill)

    add_list_validation(ws, f"F3:F{len(DEFAULT_TEAM) + 10}", ["Yes", "No"])
    set_col_widths(ws, [4, 24, 22, 20, 28, 10, 30])
    ws.freeze_panes = "A3"


def build_year_dashboard(ws):
    ws.title = "Year Dashboard"
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{YEAR} MONTHLY SYNC — YEAR AT A GLANCE"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=16, color=WHITE)

    headers = [
        "Month", "Sync Date", "Facilitator", "Theme / Focus",
        "Pre-read Sent?", "Sync Done?", "Action Items Open",
        "Key Wins", "Top Gaps", "Next Month Preview",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=SUBHEADER_FILL, align="center", size=10)

    for i, (month_name, _) in enumerate(MONTHS, 1):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 1, month_name)
        for col in range(2, 11):
            style_cell(ws.cell(row, col), fill=fill)
        style_cell(ws.cell(row, 1), fill=fill, bold=True)
        ws.cell(row, 5, "No")
        ws.cell(row, 6, "Not Started")
        style_cell(ws.cell(row, 5), fill=fill, align="center")
        style_cell(ws.cell(row, 6), fill=fill, align="center")

    last_row = len(MONTHS) + 2
    add_list_validation(ws, f"E3:E{last_row}", ["Yes", "No"])
    add_list_validation(ws, f"F3:F{last_row}", ["Not Started", "Scheduled", "Done", "Skipped"])
    add_status_colors(ws, f"F3:F{last_row}")

    set_col_widths(ws, [12, 14, 18, 28, 12, 12, 14, 28, 28, 28])
    ws.freeze_panes = "B3"
    ws.sheet_view.zoomScale = 85


def build_monthly_sheet(ws, month_name, short_name):
    ws.title = f"{short_name} {YEAR}"
    ws.merge_cells("A1:K1")
    ws["A1"] = f"{month_name.upper()} {YEAR} — MONTHLY TECH SYNC"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    # Meeting metadata
    meta = [
        ("Sync Date:", "B2"), ("Time:", "D2"), ("Duration:", "F2"), ("Location / Link:", "H2"),
        ("Facilitator:", "B3"), ("Note Taker:", "D3"), ("Attendees:", "F3"),
    ]
    labels = {
        "B2": "Sync Date:", "D2": "Time:", "F2": "Duration:", "H2": "Location / Link:",
        "B3": "Facilitator:", "D3": "Note Taker:", "F3": "Attendees:",
    }
    for cell_ref, label in labels.items():
        col = ord(cell_ref[0]) - ord("A") + 1
        row = int(cell_ref[1])
        ws.cell(row, col - 1 if col > 1 else 1, label)
        # Simpler: place labels in A2, C2, E2, G2, A3, C3, E3
    ws["A2"], ws["C2"], ws["E2"], ws["G2"] = "Sync Date:", "Time:", "Duration:", "Location / Link:"
    ws["A3"], ws["C3"], ws["E3"] = "Facilitator:", "Note Taker:", "Attendees:"
    for ref in ["A2", "C2", "E2", "G2", "A3", "C3", "E3"]:
        style_cell(ws[ref], bold=True, fill=SECTION_FILL)
    for ref in ["B2", "D2", "F2", "H2", "B3", "D3", "F3"]:
        style_cell(ws[ref], fill=WHITE_FILL)
    ws.merge_cells("F3:H3")

    # Team agenda section
    row = 5
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    ws.cell(row, 1, "MONTHLY AGENDA")
    style_cell(ws.cell(row, 1), bold=True, fill=PatternFill("solid", fgColor=BLUE), align="center", color=WHITE)
    row += 1

    agenda_headers = ["#", "Agenda Item", "Owner", "Time (min)", "Status", "Notes"]
    agenda_cols = [1, 2, 5, 7, 8, 9]
    for col_idx, h in zip(agenda_cols, agenda_headers):
        ws.cell(row, col_idx, h)
        style_cell(ws.cell(row, col_idx), bold=True, fill=SUBHEADER_FILL, align="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
    row += 1

    default_agenda = [
        (1, "Review last month's action items", "Manager", 5),
        (2, "Member round-robin updates", "All", 20),
        (3, "Deep dive: top blockers & dependencies", "Tech Lead", 15),
        (4, "Feedback & process improvements", "Manager", 10),
        (5, "Confirm action items & next month preview", "Manager", 10),
    ]
    agenda_start = row
    for num, item, owner, mins in default_agenda:
        ws.cell(row, 1, num)
        ws.cell(row, 2, item)
        ws.cell(row, 5, owner)
        ws.cell(row, 7, mins)
        ws.cell(row, 8, "Not Started")
        fill = ALT_FILL if num % 2 == 0 else WHITE_FILL
        for c in [1, 2, 5, 7, 8, 9]:
            style_cell(ws.cell(row, c), fill=fill, align="center" if c in (1, 7, 8) else "left")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
        row += 1
    agenda_end = row - 1
    add_list_validation(ws, f"H{agenda_start}:H{agenda_end}", ["Not Started", "In Progress", "Done", "Skipped"])
    add_status_colors(ws, f"H{agenda_start}:H{agenda_end}")

    # Blank agenda rows
    for num in range(6, 9):
        ws.cell(row, 1, num)
        fill = ALT_FILL if num % 2 == 0 else WHITE_FILL
        for c in [1, 2, 5, 7, 8, 9]:
            style_cell(ws.cell(row, c), fill=fill)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    ws.cell(row, 1, "MEMBER UPDATES — Each person fills their row before the sync")
    style_cell(ws.cell(row, 1), bold=True, fill=PatternFill("solid", fgColor=TEAL), align="center", color=WHITE)
    row += 1

    member_headers = [
        "Name", "Role", "Last Month Highlights / Updates",
        "New Items / Learnings / Deliverables", "What's Next (30 days)",
        "Gaps / Blockers / Where We Need Help", "Feedback Needed (from Manager/Team)",
        "Support Requested", "Confidence (1-5)", "Status", "Manager Notes",
    ]
    for col, h in enumerate(member_headers, 1):
        ws.cell(row, col, h)
        style_cell(ws.cell(row, col), bold=True, fill=SUBHEADER_FILL, align="center", size=10)
    row += 1

    member_start = row
    for i, (name, role, squad) in enumerate(DEFAULT_TEAM, 1):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 1, name)
        ws.cell(row, 2, role)
        ws.cell(row, 10, "Not Started")
        for col in range(1, 12):
            style_cell(ws.cell(row, col), fill=fill, align="center" if col in (9, 10) else "left", size=10)
        row += 1

    # Extra blank member rows
    for i in range(len(DEFAULT_TEAM) + 1, len(DEFAULT_TEAM) + 5):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 10, "Not Started")
        for col in range(1, 12):
            style_cell(ws.cell(row, col), fill=fill, size=10)
        row += 1
    member_end = row - 1

    add_list_validation(ws, f"J{member_start}:J{member_end}", ["Not Started", "In Progress", "Done", "Blocked"])
    add_list_validation(ws, f"I{member_start}:I{member_end}", ["1", "2", "3", "4", "5"])
    add_status_colors(ws, f"J{member_start}:J{member_end}")

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    ws.cell(row, 1, "MONTH SUMMARY & DECISIONS")
    style_cell(ws.cell(row, 1), bold=True, fill=SUBHEADER_FILL, align="center")
    row += 1

    summary_fields = [
        "Key Wins This Month:",
        "Top 3 Gaps to Address:",
        "Decisions Made:",
        "Carry-over to Next Month:",
        "Draft Next Month Agenda:",
    ]
    for field in summary_fields:
        ws.cell(row, 1, field)
        style_cell(ws.cell(row, 1), bold=True, fill=SECTION_FILL)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        style_cell(ws.cell(row, 2), fill=WHITE_FILL)
        ws.row_dimensions[row].height = 36
        row += 1

    set_col_widths(ws, [18, 16, 28, 28, 24, 26, 24, 18, 12, 12, 24])
    ws.freeze_panes = f"A{member_start}"
    ws.sheet_view.zoomScale = 80


def build_action_items_sheet(ws):
    ws.title = "Action Items Tracker"
    ws.merge_cells("A1:J1")
    ws["A1"] = "CROSS-MONTH ACTION ITEMS TRACKER"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    headers = [
        "ID", "Month Raised", "Action Item", "Owner", "Priority",
        "Due Date", "Status", "Blocked By", "Completed Date", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=SUBHEADER_FILL, align="center", size=10)

    for i in range(1, 51):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 1, f"AI-{i:03d}")
        ws.cell(row, 7, "Not Started")
        for col in range(1, 11):
            style_cell(ws.cell(row, col), fill=fill, align="center" if col in (1, 5, 7) else "left", size=10)

    add_list_validation(ws, "E3:E52", ["High", "Medium", "Low"])
    add_list_validation(ws, "G3:G52", ["Not Started", "In Progress", "Blocked", "Done", "Cancelled"])
    add_status_colors(ws, "G3:G52")
    set_col_widths(ws, [8, 12, 36, 18, 10, 12, 12, 20, 14, 28])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:J52"


def build_feedback_log_sheet(ws):
    ws.title = "Feedback Log"
    ws.merge_cells("A1:H1")
    ws["A1"] = "TEAM FEEDBACK LOG"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    headers = [
        "Date", "Month", "From (Name)", "Category", "Feedback / Concern",
        "Manager Response", "Action Taken", "Status",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=SUBHEADER_FILL, align="center", size=10)

    categories = [
        "Process", "Tools", "Communication", "Workload", "Career Growth",
        "Team Culture", "Technical Debt", "Other",
    ]
    month_options = [m[0] for m in MONTHS]

    for i in range(1, 41):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 8, "Open")
        for col in range(1, 9):
            style_cell(ws.cell(row, col), fill=fill, size=10)

    add_list_validation(ws, "B3:B42", month_options)
    add_list_validation(ws, "D3:D42", categories)
    add_list_validation(ws, "H3:H42", ["Open", "In Progress", "Resolved", "Won't Fix"])
    add_status_colors(ws, "H3:H42")
    set_col_widths(ws, [12, 12, 18, 16, 36, 28, 24, 12])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:H42"


def build_gap_tracker_sheet(ws):
    ws.title = "Gap & Growth Tracker"
    ws.merge_cells("A1:I1")
    ws["A1"] = "SKILLS, PROCESS & RESOURCE GAPS"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    headers = [
        "Gap ID", "Area", "Gap Description", "Impact", "Affected Members",
        "Proposed Solution", "Owner", "Target Month", "Status",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=SUBHEADER_FILL, align="center", size=10)

    areas = ["Skills", "Process", "Tools", "Documentation", "Hiring", "Knowledge", "Other"]
    month_options = [m[0] for m in MONTHS]
    impacts = ["High", "Medium", "Low"]

    for i in range(1, 31):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 1, f"GAP-{i:03d}")
        ws.cell(row, 9, "Open")
        for col in range(1, 10):
            style_cell(ws.cell(row, col), fill=fill, size=10)

    add_list_validation(ws, "B3:B32", areas)
    add_list_validation(ws, "D3:D32", impacts)
    add_list_validation(ws, "H3:H32", month_options)
    add_list_validation(ws, "I3:I32", ["Open", "In Progress", "Resolved", "Deferred"])
    add_status_colors(ws, "I3:I32")
    set_col_widths(ws, [8, 14, 32, 10, 22, 28, 16, 14, 12])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:I32"


def build_agenda_library_sheet(ws):
    ws.title = "Agenda Library"
    ws.merge_cells("A1:E1")
    ws["A1"] = "REUSABLE AGENDA ITEMS FOR MONTHLY SYNCS"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    headers = ["#", "Agenda Item", "Typical Duration", "When to Use", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=SUBHEADER_FILL, align="center")

    library = [
        ("Review last month's action items", "5 min", "Every sync", "Start every meeting here"),
        ("Member round-robin updates", "20 min", "Every sync", "2-3 min per person"),
        ("Sprint / release retrospective", "15 min", "End of sprint month", "What went well / didn't"),
        ("Technical debt review", "15 min", "Quarterly", "Prioritize paydown items"),
        ("Architecture / design discussion", "20 min", "When planning features", "Align before big builds"),
        ("Security & compliance check-in", "10 min", "Quarterly", "Vulnerabilities, audits"),
        ("Career growth & learning goals", "15 min", "Bi-monthly", "1:1 topics rolled into team sync"),
        ("Cross-team dependency mapping", "15 min", "When blocked", "Identify handoffs"),
        ("Process improvement brainstorm", "15 min", "When friction reported", "Start/stop/continue"),
        ("OKR / goal progress review", "15 min", "Quarter start/end", "Align with company goals"),
        ("Incident post-mortem (if any)", "20 min", "After incidents", "Blameless review"),
        ("Hiring & capacity planning", "15 min", "When hiring", "Workload vs headcount"),
        ("Tooling & dev environment feedback", "10 min", "Bi-monthly", "CI/CD, IDE, access issues"),
        ("Knowledge sharing session", "20 min", "Monthly rotation", "One person presents"),
        ("Preview next month priorities", "10 min", "Close every sync", "Set expectations early"),
    ]

    for i, (item, duration, when, notes) in enumerate(library, 1):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row, 1, i)
        ws.cell(row, 2, item)
        ws.cell(row, 3, duration)
        ws.cell(row, 4, when)
        ws.cell(row, 5, notes)
        for col in range(1, 6):
            style_cell(ws.cell(row, col), fill=fill)

    set_col_widths(ws, [4, 36, 14, 22, 32])
    ws.freeze_panes = "A3"


def main():
    wb = Workbook()

    ws_start = wb.active
    build_instructions_sheet(ws_start)

    ws_roster = wb.create_sheet("Team Roster")
    build_roster_sheet(ws_roster)

    ws_dashboard = wb.create_sheet("Year Dashboard")
    build_year_dashboard(ws_dashboard)

    for month_name, short_name in MONTHS:
        ws = wb.create_sheet(f"{short_name} {YEAR}")
        build_monthly_sheet(ws, month_name, short_name)

    ws_actions = wb.create_sheet("Action Items Tracker")
    build_action_items_sheet(ws_actions)

    ws_feedback = wb.create_sheet("Feedback Log")
    build_feedback_log_sheet(ws_feedback)

    ws_gaps = wb.create_sheet("Gap & Growth Tracker")
    build_gap_tracker_sheet(ws_gaps)

    ws_agenda = wb.create_sheet("Agenda Library")
    build_agenda_library_sheet(ws_agenda)

    output = "tech_team_monthly_sync_2027.xlsx"
    wb.save(output)
    print(f"Created {output}")


if __name__ == "__main__":
    main()
