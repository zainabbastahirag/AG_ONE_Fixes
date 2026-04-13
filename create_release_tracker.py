#!/usr/bin/env python3
"""
AG ONE — Release History Tracker (Separate Excel)
Dedicated release management workbook: Dashboard + per-product tabs + standard
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ─── palette ──────────────────────────────────────────────────────────
NAVY       = "1E3A5F"
BLUE       = "3B82F6"
LIGHT_BLUE = "DBEAFE"
WHITE      = "FFFFFF"
DARK       = "374151"
GRAY       = "6B7280"
LGRAY      = "F3F4F6"
BORDER     = "E2E8F0"
GREEN      = "10B981"
LGREEN     = "D1FAE5"
AMBER      = "F59E0B"
LAMBER     = "FEF3C7"
RED        = "EF4444"
LRED       = "FEE2E2"
SKY        = "0EA5E9"
LSKY       = "E0F2FE"

thin = Border(
    left=Side("thin", BORDER), right=Side("thin", BORDER),
    top=Side("thin", BORDER), bottom=Side("thin", BORDER),
)

PRODUCTS = {
    "AG ONE":  {"color": "1E3A5F", "light": "DBEAFE", "lead": "Geena"},
    "OneWork": {"color": "3B82F6", "light": "DBEAFE", "lead": "Abdullah"},
    "Learn":   {"color": "8B5CF6", "light": "EDE9FE", "lead": "Ricky"},
    "Safe":    {"color": "10B981", "light": "D1FAE5", "lead": "Faisal"},
    "OneHire": {"color": "F59E0B", "light": "FEF3C7", "lead": "Logesh"},
    "Spot":    {"color": "EF4444", "light": "FEE2E2", "lead": "Ricky"},
    "Pulse":   {"color": "EC4899", "light": "FCE7F3", "lead": "Hanis"},
}

TYPE_STYLE = {
    "MVP":    ("10B981", "D1FAE5", "065F46"),
    "Major":  ("1E3A5F", "DBEAFE", "1E3A5F"),
    "Minor":  ("3B82F6", "DBEAFE", "1E3A5F"),
    "Hotfix": ("EF4444", "FEE2E2", "991B1B"),
    "Patch":  ("F59E0B", "FEF3C7", "92400E"),
    "Sprint": ("8B5CF6", "EDE9FE", "5B21B6"),
}

RELEASES = [
    # AG ONE
    {"product":"AG ONE","version":"v1.0.0","name":"AG ONE MVP","type":"MVP","date":"29 Mar 2026","sprint":"Sprint 6","status":"✅ Released","highlights":"Full IAM platform: Users, Roles, Tenants, SSO, Permissions, Audit, Assign Access"},
    {"product":"AG ONE","version":"v1.0.1","name":"AG ONE Hotfix 1","type":"Hotfix","date":"02 Apr 2026","sprint":"Sprint 7","status":"✅ Released","highlights":"SSO token refresh loop fix, role sync cache invalidation, tenant switching fix"},
    {"product":"AG ONE","version":"v1.0.2","name":"AG ONE Hotfix 2","type":"Hotfix","date":"08 Apr 2026","sprint":"Sprint 7","status":"✅ Released","highlights":"Audit log pagination, API key duplicate name validation, timezone fix"},
    {"product":"AG ONE","version":"v1.1.0","name":"AG ONE Q2 Enhancement","type":"Minor","date":"May 2026","sprint":"Sprint 9–10","status":"📋 Planned","highlights":"Advanced user analytics, bulk role operations, enhanced audit dashboard"},
    {"product":"AG ONE","version":"v1.2.0","name":"AG ONE Multi-Product Admin","type":"Minor","date":"Jul 2026","sprint":"Sprint 13–14","status":"📋 Planned","highlights":"Multi-product admin console, cross-product role mapping, API marketplace v1"},
    {"product":"AG ONE","version":"v2.0.0","name":"AG ONE Enterprise","type":"Major","date":"Oct 2026","sprint":"Sprint 20–22","status":"📋 Planned","highlights":"Enterprise SSO (SAML+OIDC), advanced RBAC, white-label branding, scale"},
    # OneWork
    {"product":"OneWork","version":"v1.0.0","name":"OneWork MVP","type":"MVP","date":"29 Mar 2026","sprint":"Sprint 6","status":"✅ Released","highlights":"Employee mgmt, workflows, task mgmt, dashboards, reporting"},
    {"product":"OneWork","version":"v1.1.0","name":"OneWork Q2 Enhancement","type":"Minor","date":"Jun 2026","sprint":"Sprint 11–12","status":"📋 Planned","highlights":"Advanced workflow templates, email notifications, calendar integration"},
    {"product":"OneWork","version":"v1.2.0","name":"OneWork Reporting v2","type":"Minor","date":"Aug 2026","sprint":"Sprint 15–16","status":"📋 Planned","highlights":"Reporting v2 with charts, mobile-responsive, third-party integrations"},
    {"product":"OneWork","version":"v2.0.0","name":"OneWork Enterprise","type":"Major","date":"Nov 2026","sprint":"Sprint 21–23","status":"📋 Planned","highlights":"AI-powered insights, advanced automation, enterprise scale"},
    # Learn
    {"product":"Learn","version":"v1.0.0","name":"Learn MVP","type":"MVP","date":"29 Mar 2026","sprint":"Sprint 6","status":"✅ Released","highlights":"Courses, learning paths, assessments, certifications, progress tracking"},
    {"product":"Learn","version":"v1.1.0","name":"Learn Analytics","type":"Minor","date":"Jun 2026","sprint":"Sprint 11–12","status":"📋 Planned","highlights":"Advanced analytics dashboard, learner recommendations, completion reports"},
    {"product":"Learn","version":"v1.2.0","name":"Learn Marketplace","type":"Minor","date":"Aug 2026","sprint":"Sprint 15–16","status":"📋 Planned","highlights":"Content marketplace, SCORM support, mobile-responsive learning"},
    {"product":"Learn","version":"v2.0.0","name":"Learn Enterprise","type":"Major","date":"Nov 2026","sprint":"Sprint 21–23","status":"📋 Planned","highlights":"AI recommendations, adaptive learning, full mobile app"},
    # Safe
    {"product":"Safe","version":"v1.0.0","name":"Safe MVP","type":"MVP","date":"29 Mar 2026","sprint":"Sprint 6","status":"✅ Released","highlights":"Policies, compliance tracking, audit workflows, risk matrix, dashboards"},
    {"product":"Safe","version":"v1.1.0","name":"Safe Regulatory Auto","type":"Minor","date":"Jun 2026","sprint":"Sprint 11–12","status":"📋 Planned","highlights":"Regulatory update automation, scheduled compliance checks, advanced reporting"},
    {"product":"Safe","version":"v1.2.0","name":"Safe Integrations","type":"Minor","date":"Sep 2026","sprint":"Sprint 17–18","status":"📋 Planned","highlights":"Third-party GRC integrations, automated alerts, evidence collection"},
    {"product":"Safe","version":"v2.0.0","name":"Safe Enterprise","type":"Major","date":"Dec 2026","sprint":"Sprint 24–26","status":"📋 Planned","highlights":"Enterprise compliance suite with AI-driven risk predictions"},
    # OneHire
    {"product":"OneHire","version":"v0.1.0","name":"OneHire Alpha","type":"Sprint","date":"May 2026","sprint":"Sprint 9–10","status":"📋 Planned","highlights":"Core job posting and applicant tracking features"},
    {"product":"OneHire","version":"v0.5.0","name":"OneHire Beta","type":"Sprint","date":"Jun 2026","sprint":"Sprint 11–12","status":"📋 Planned","highlights":"Interview scheduling, candidate pipeline, basic reporting"},
    {"product":"OneHire","version":"v1.0.0","name":"OneHire MVP","type":"MVP","date":"Jul 2026","sprint":"Sprint 13–14","status":"📋 Planned","highlights":"Full recruitment portal with applicant tracking and analytics"},
    # Spot
    {"product":"Spot","version":"v0.1.0","name":"Spot Alpha","type":"Sprint","date":"May 2026","sprint":"Sprint 9–10","status":"📋 Planned","highlights":"Core city module, location management"},
    {"product":"Spot","version":"v1.0.0","name":"Spot MVP","type":"MVP","date":"Sep 2026","sprint":"Sprint 17–18","status":"📋 Planned","highlights":"City management, geolocation features, admin dashboard"},
    # Pulse
    {"product":"Pulse","version":"v0.1.0","name":"Pulse Alpha","type":"Sprint","date":"May 2026","sprint":"Sprint 9–10","status":"📋 Planned","highlights":"Survey builder, basic distribution"},
    {"product":"Pulse","version":"v1.0.0","name":"Pulse MVP","type":"MVP","date":"Sep 2026","sprint":"Sprint 17–18","status":"📋 Planned","highlights":"Survey analytics, employee engagement dashboards, benchmarking"},
]


# ═══════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════

def hdr(ws, row, cols, color=NAVY):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

def title(ws, row, text, sub, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Aptos", bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for x in range(2, ncols+1):
        ws.cell(row=row, column=x).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[row].height = 36

    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
    s = ws.cell(row=row+1, column=1, value=sub)
    s.font = Font(name="Aptos", size=10, italic=True, color=GRAY)
    s.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    s.alignment = Alignment(horizontal="left", vertical="center")
    for x in range(2, ncols+1):
        ws.cell(row=row+1, column=x).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.row_dimensions[row+1].height = 22

def dcell(cell, row, center=True):
    fill = LGRAY if row % 2 == 0 else WHITE
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=10, color=DARK)
    cell.border = thin
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)

def banner(ws, row, text, bg, fg, ncols, height=30):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Aptos", bold=True, size=11, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin
    for x in range(2, ncols+1):
        ws.cell(row=row, column=x).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=row, column=x).border = thin
    ws.row_dimensions[row].height = height

def aw(ws, mn=10, mx=32):
    for col in ws.columns:
        ml = 0
        cl = get_column_letter(col[0].column)
        for c in col:
            if c.value: ml = max(ml, len(str(c.value)))
        ws.column_dimensions[cl].width = min(max(ml + 3, mn), mx)


# ═══════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ───────────────────────────────────────────────────────────────────────
# SHEET 1 — RELEASE DASHBOARD
# ───────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Release Dashboard"
ws1.sheet_properties.tabColor = SKY

COLS1 = ["Product", "Tech Lead", "Latest Version", "Latest Release", "Release Date",
         "Total Released", "Hotfixes", "Next Planned", "Next Version", "Target Date"]
nc1 = len(COLS1)
title(ws1, 1, "AG ONE — Release Dashboard 2026",
      f"All products at a glance  •  Last updated: {date.today().strftime('%d %b %Y')}", nc1)

banner(ws1, 3, "📏  Standard: vMAJOR.MINOR.PATCH  •  Types: MVP | Major | Minor | Hotfix | Patch | Sprint  •  Quarterly + hotfixes as needed",
       LSKY, "0369A1", nc1, 28)

hdr(ws1, 4, COLS1, SKY)

r = 5
for prod, info in PRODUCTS.items():
    tc = info["color"]
    rels = [x for x in RELEASES if x["product"] == prod]
    done = [x for x in rels if "Released" in x["status"]]
    hf = [x for x in done if x["type"] == "Hotfix"]
    plan = [x for x in rels if "Planned" in x["status"]]
    latest = done[-1] if done else None
    nxt = plan[0] if plan else None

    ws1.cell(row=r, column=1, value=prod)
    ws1.cell(row=r, column=1).font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=tc)
    ws1.cell(row=r, column=2, value=info["lead"])
    ws1.cell(row=r, column=3, value=latest["version"] if latest else "—")
    ws1.cell(row=r, column=4, value=latest["name"] if latest else "—")
    ws1.cell(row=r, column=5, value=latest["date"] if latest else "—")
    ws1.cell(row=r, column=6, value=len(done))
    ws1.cell(row=r, column=7, value=len(hf))
    ws1.cell(row=r, column=8, value=nxt["name"] if nxt else "—")
    ws1.cell(row=r, column=9, value=nxt["version"] if nxt else "—")
    ws1.cell(row=r, column=10, value=nxt["date"] if nxt else "—")

    for c in range(1, nc1+1):
        cell = ws1.cell(row=r, column=c)
        if c != 1: dcell(cell, r)
        else:
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    if len(done) > 0:
        ws1.cell(row=r, column=6).font = Font(name="Aptos", bold=True, size=11, color="065F46")
        ws1.cell(row=r, column=6).fill = PatternFill("solid", fgColor=LGREEN)
    if len(hf) > 0:
        ws1.cell(row=r, column=7).font = Font(name="Aptos", bold=True, size=10, color="991B1B")
        ws1.cell(row=r, column=7).fill = PatternFill("solid", fgColor=LRED)
    ws1.row_dimensions[r].height = 32
    r += 1

# Totals
r += 1
tot_done = len([x for x in RELEASES if "Released" in x["status"]])
tot_hf = len([x for x in RELEASES if "Released" in x["status"] and x["type"] == "Hotfix"])
tot_plan = len([x for x in RELEASES if "Planned" in x["status"]])
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws1.cell(row=r, column=1, value="TOTAL ACROSS ALL PRODUCTS")
ws1.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=11, color=WHITE)
ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
for c in range(2, 6):
    ws1.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NAVY)
ws1.cell(row=r, column=6, value=tot_done)
ws1.cell(row=r, column=6).font = Font(name="Aptos", bold=True, size=12, color="065F46")
ws1.cell(row=r, column=6).fill = PatternFill("solid", fgColor=LGREEN)
ws1.cell(row=r, column=7, value=tot_hf)
ws1.cell(row=r, column=7).font = Font(name="Aptos", bold=True, size=12, color="991B1B")
ws1.cell(row=r, column=7).fill = PatternFill("solid", fgColor=LRED)
ws1.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
ws1.cell(row=r, column=8, value=f"{tot_plan} releases planned")
ws1.cell(row=r, column=8).font = Font(name="Aptos", bold=True, size=11, color=GRAY)
for c in range(6, nc1+1):
    ws1.cell(row=r, column=c).border = thin
    ws1.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

ws1.freeze_panes = "B5"
aw(ws1)
ws1.column_dimensions["D"].width = 28
ws1.column_dimensions["H"].width = 28


# ───────────────────────────────────────────────────────────────────────
# SHEETS 2–8 — PER-PRODUCT RELEASE TABS
# ───────────────────────────────────────────────────────────────────────
COLS_P = ["#", "Version", "Release Name", "Type", "Date", "Sprint", "Status", "What's Included / Highlights"]
ncp = len(COLS_P)

for prod, info in PRODUCTS.items():
    tc = info["color"]
    tl = info["light"]
    rels = [x for x in RELEASES if x["product"] == prod]
    done = [x for x in rels if "Released" in x["status"]]
    plan = [x for x in rels if "Planned" in x["status"]]
    hf = [x for x in done if x["type"] == "Hotfix"]

    ws = wb.create_sheet(prod)
    ws.sheet_properties.tabColor = tc

    title(ws, 1, f"{prod} — Release History & Roadmap",
          f"Tech Lead: {info['lead']}  •  Released: {len(done)}  •  Hotfixes: {len(hf)}  •  Planned: {len(plan)}", ncp)

    # Stats banner
    latest_v = done[-1]["version"] if done else "—"
    next_v = f"{plan[0]['version']} ({plan[0]['date']})" if plan else "—"
    banner(ws, 3, f"Latest: {latest_v}  •  Next: {next_v}  •  Total: {len(rels)} releases",
           tl, tc, ncp, 28)

    # ── RELEASED ──
    r = 4
    banner(ws, r, "✅  RELEASED", LGREEN, "065F46", ncp, 26)
    r = 5
    hdr(ws, r, COLS_P, tc)
    r = 6

    if done:
        for idx, rel in enumerate(done, 1):
            ts = TYPE_STYLE.get(rel["type"], ("6B7280", LGRAY, DARK))
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=rel["version"])
            ws.cell(row=r, column=3, value=rel["name"])
            ws.cell(row=r, column=4, value=rel["type"])
            ws.cell(row=r, column=5, value=rel["date"])
            ws.cell(row=r, column=6, value=rel["sprint"])
            ws.cell(row=r, column=7, value=rel["status"])
            ws.cell(row=r, column=8, value=rel["highlights"])

            for c in range(1, ncp+1):
                cell = ws.cell(row=r, column=c)
                dcell(cell, r, center=(c not in [3, 8]))
            ws.cell(row=r, column=2).font = Font(name="Aptos", bold=True, size=11, color=DARK)
            ws.cell(row=r, column=4).font = Font(name="Aptos", bold=True, size=10, color=ts[2])
            ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=ts[1])
            ws.cell(row=r, column=7).font = Font(name="Aptos", bold=True, size=10, color="065F46")
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor=LGREEN)
            ws.row_dimensions[r].height = 34
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncp)
        ws.cell(row=r, column=1, value="No releases yet — development in progress")
        ws.cell(row=r, column=1).font = Font(name="Aptos", italic=True, size=10, color=GRAY)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=1).border = thin
        for c in range(2, ncp+1): ws.cell(row=r, column=c).border = thin
        r += 1

    # Blank rows for future released entries
    for _ in range(3):
        for c in range(1, ncp+1):
            dcell(ws.cell(row=r, column=c), r, center=(c not in [3, 8]))
        r += 1

    # ── UPCOMING ──
    r += 1
    banner(ws, r, "📋  UPCOMING / PLANNED", LSKY, "0369A1", ncp, 26)
    r += 1
    hdr(ws, r, COLS_P, "0284C7")
    r += 1

    if plan:
        for idx, rel in enumerate(plan, 1):
            ts = TYPE_STYLE.get(rel["type"], ("6B7280", LGRAY, DARK))
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=rel["version"])
            ws.cell(row=r, column=3, value=rel["name"])
            ws.cell(row=r, column=4, value=rel["type"])
            ws.cell(row=r, column=5, value=rel["date"])
            ws.cell(row=r, column=6, value=rel["sprint"])
            ws.cell(row=r, column=7, value=rel["status"])
            ws.cell(row=r, column=8, value=rel["highlights"])

            for c in range(1, ncp+1):
                cell = ws.cell(row=r, column=c)
                dcell(cell, r, center=(c not in [3, 8]))
            ws.cell(row=r, column=2).font = Font(name="Aptos", bold=True, size=11, color=DARK)
            ws.cell(row=r, column=4).font = Font(name="Aptos", bold=True, size=10, color=ts[2])
            ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=ts[1])
            ws.cell(row=r, column=7).font = Font(name="Aptos", italic=True, size=10, color=GRAY)
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor=LGRAY)
            ws.row_dimensions[r].height = 34
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncp)
        ws.cell(row=r, column=1, value="No upcoming releases planned yet")
        ws.cell(row=r, column=1).font = Font(name="Aptos", italic=True, size=10, color=GRAY)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=1).border = thin
        for c in range(2, ncp+1): ws.cell(row=r, column=c).border = thin
        r += 1

    # Blank rows for future planned entries
    for _ in range(3):
        for c in range(1, ncp+1):
            dcell(ws.cell(row=r, column=c), r, center=(c not in [3, 8]))
        r += 1

    ws.freeze_panes = "C6"
    aw(ws)
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["H"].width = 55


# ───────────────────────────────────────────────────────────────────────
# SHEET 9 — RELEASE NAMING STANDARD
# ───────────────────────────────────────────────────────────────────────
ws_s = wb.create_sheet("Release Standard")
ws_s.sheet_properties.tabColor = "0369A1"
nc_s = 4
title(ws_s, 1, "Release Naming Standard & Versioning Policy",
      "Share with all teams — standard format for naming and numbering releases", nc_s)

r = 3
hdr(ws_s, r, ["Component", "Description", "Example", "Notes"], NAVY)
r = 4
fmt = [
    ("Format", "vMAJOR.MINOR.PATCH", "v1.2.3", "Semantic versioning"),
    ("MAJOR (X.0.0)", "Breaking changes, major feature sets", "v1.0.0 → v2.0.0", "Needs Director approval"),
    ("MINOR (1.X.0)", "New features, backward compatible", "v1.0.0 → v1.1.0", "Quarterly cadence"),
    ("PATCH (1.0.X)", "Bug fixes, hotfixes", "v1.0.0 → v1.0.1", "As needed"),
]
for label, desc, ex, note in fmt:
    ws_s.cell(row=r, column=1, value=label)
    ws_s.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=11, color=WHITE)
    ws_s.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws_s.cell(row=r, column=2, value=desc)
    ws_s.cell(row=r, column=3, value=ex)
    ws_s.cell(row=r, column=4, value=note)
    for c in range(1, nc_s+1):
        cell = ws_s.cell(row=r, column=c)
        if c > 1: dcell(cell, r, center=False)
        else:
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

r += 1
hdr(ws_s, r, ["Release Type", "Description", "Frequency", "Approval Required"], SKY)
r += 1
types = [
    ("MVP", "First production release", "Once per product", "Director + Tech Lead", "10B981"),
    ("Major", "Significant new features or breaking changes", "1–2 per year", "Director + Tech Lead", "1E3A5F"),
    ("Minor", "New features, enhancements, backward compatible", "Quarterly", "Tech Lead", "3B82F6"),
    ("Hotfix", "Critical production fix, expedited", "As needed (urgent)", "Tech Lead + post-mortem", "EF4444"),
    ("Patch", "Non-critical bug fixes", "As needed", "Tech Lead", "F59E0B"),
    ("Sprint", "Pre-MVP iteration (alpha/beta)", "Every 2 weeks", "Tech Lead", "8B5CF6"),
]
for tp, desc, freq, appr, clr in types:
    ws_s.cell(row=r, column=1, value=tp)
    ws_s.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=10, color=WHITE)
    ws_s.cell(row=r, column=1).fill = PatternFill("solid", fgColor=clr)
    ws_s.cell(row=r, column=2, value=desc)
    ws_s.cell(row=r, column=3, value=freq)
    ws_s.cell(row=r, column=4, value=appr)
    for c in range(1, nc_s+1):
        cell = ws_s.cell(row=r, column=c)
        if c > 1: dcell(cell, r, center=False)
        else:
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

r += 1
hdr(ws_s, r, ["Naming Pattern", "Example", "When to Use", ""], NAVY)
r += 1
names = [
    ("[Product] MVP", "AG ONE MVP", "First production release"),
    ("[Product] Q[N] Enhancement", "AG ONE Q2 Enhancement", "Quarterly minor release"),
    ("[Product] Hotfix [N]", "AG ONE Hotfix 1", "Emergency fix, numbered"),
    ("[Product] [Theme]", "AG ONE Enterprise", "Major themed release"),
    ("[Product] Alpha / Beta", "OneHire Alpha", "Pre-MVP sprint iterations"),
]
for pat, ex, when in names:
    ws_s.cell(row=r, column=1, value=pat)
    ws_s.cell(row=r, column=1).font = Font(name="Aptos", bold=True, size=10, color=NAVY)
    ws_s.cell(row=r, column=2, value=ex)
    ws_s.cell(row=r, column=2).font = Font(name="Aptos", bold=True, size=10, color=DARK)
    ws_s.cell(row=r, column=3, value=when)
    for c in range(1, nc_s+1):
        cell = ws_s.cell(row=r, column=c)
        if c != 1: dcell(cell, r, center=False)
        else:
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
    r += 1

aw(ws_s)
ws_s.column_dimensions["A"].width = 26
ws_s.column_dimensions["B"].width = 44
ws_s.column_dimensions["C"].width = 28
ws_s.column_dimensions["D"].width = 28


# ─── FINAL ────────────────────────────────────────────────────────────
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

out = "/workspace/AG_ONE_Release_History.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
print(f"   Sheets: {wb.sheetnames}")
