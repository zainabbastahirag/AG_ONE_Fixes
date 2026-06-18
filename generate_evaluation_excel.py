#!/usr/bin/env python3
"""Generate candidate evaluation Excel workbook."""

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PHASE1_CANDIDATES = [
    "Mohammad Kashif Mohiuddin",
    "Shahid",
    "Hadeef",
]

PHASE2_CANDIDATES = [
    "Baizid Yaldram",
    "MUHAMMAD HADIF AMAR BIN RAZALI",
    "CITA WAFA ATIAH",
    "MUHAMMAD MAIRAJ",
    "Mohammad Shahid Akhtar",
    "Aurneela Ghosh Riddhi",
]

PHASE1_CRITERIA = [
    ("Presentation", "Structure, confidence, slides, delivery"),
    ("Technical Observation", "Spots and explains technical points well"),
]

PHASE2_CRITERIA = [
    ("Clear Thinking", "Clarity of thought — easy to follow"),
    ("Deep vs Wide", "Depth vs breadth of answers"),
    ("Build on Ideas", "Listing and building on points"),
    ("Business Link", "Links ideas to business value"),
    ("Positive Manner", "Conduct with benefit — respectful & calm"),
    ("Fresh Ideas", "Original thinking"),
    ("Speaking Style", "Communication style"),
]

RATING_SCALE = "1 = Weak  |  2 = Below avg  |  3 = Average  |  4 = Good  |  5 = Excellent"

NAVY = "1F4E79"
BLUE = "2E75B6"
LIGHT_BLUE = "D9E2F3"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"
ALT = "F2F2F2"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
P1_FILL = PatternFill("solid", fgColor="DDEBF7")
P2_FILL = PatternFill("solid", fgColor="E2DFEC")
ALT_FILL = PatternFill("solid", fgColor=ALT)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_cell(cell, *, bold=False, fill=None, align="left", size=11, color="000000"):
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    if fill:
        cell.fill = fill


def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_score_validation(ws, first_row, last_row, first_col, last_col):
    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv.error = "Enter a score from 1 to 5"
    dv.errorTitle = "Invalid score"
    ws.add_data_validation(dv)
    for row in range(first_row, last_row + 1):
        for col in range(first_col, last_col + 1):
            dv.add(ws.cell(row, col))


def add_score_colors(ws, cell_range):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["4"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=["3"], fill=PatternFill("solid", fgColor=YELLOW)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThanOrEqual", formula=["2"], fill=PatternFill("solid", fgColor=RED)),
    )


def build_dashboard_sheet(wb):
    ws = wb.active
    ws.title = "Dashboard Index"
    ws.merge_cells("A1:H1")
    ws["A1"] = "CANDIDATE EVALUATION — DASHBOARD INDEX"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=18, color=WHITE)

    ws.merge_cells("A2:H2")
    ws["A2"] = RATING_SCALE
    style_cell(ws["A2"], fill=SUBHEADER_FILL, align="center", size=10)

    # Session 1 block
    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "SESSION 1 — PHASE 1  |  Presentation + Technical Observation")
    style_cell(ws.cell(r, 1), bold=True, fill=PatternFill("solid", fgColor=BLUE), align="center", color=WHITE)
    r += 1

    p1_headers = ["#", "Candidate", "Presentation", "Technical Observation", "Total", "Average", "Strengths", "Notes"]
    for c, h in enumerate(p1_headers, 1):
        ws.cell(r, c, h)
        style_cell(ws.cell(r, c), bold=True, fill=P1_FILL, align="center")
    r += 1

    p1_start = r
    for i, name in enumerate(PHASE1_CANDIDATES, 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, name)
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for c in range(1, 9):
            style_cell(ws.cell(r, c), fill=fill, align="center" if c != 2 and c < 7 else "left")
        ws.cell(r, 5, f"=SUM(C{r}:D{r})")
        ws.cell(r, 6, f'=IF(COUNT(C{r}:D{r})>0,AVERAGE(C{r}:D{r}),"")')
        style_cell(ws.cell(r, 5), fill=fill, align="center", bold=True)
        style_cell(ws.cell(r, 6), fill=fill, align="center", bold=True)
        r += 1
    p1_end = r - 1
    add_score_validation(ws, p1_start, p1_end, 3, 4)
    add_score_colors(ws, f"C{p1_start}:D{p1_end}")

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "SESSION 2 — PHASE 2  |  Clarity & Discussion (simple labels)")
    style_cell(ws.cell(r, 1), bold=True, fill=PatternFill("solid", fgColor="7030A0"), align="center", color=WHITE)
    r += 1

    p2_headers = [
        "#", "Candidate", "Clear Thinking", "Deep vs Wide", "Build on Ideas",
        "Business Link", "Positive Manner", "Fresh Ideas", "Speaking Style",
        "Total", "Average", "Strengths", "Notes",
    ]
    for c, h in enumerate(p2_headers, 1):
        ws.cell(r, c, h)
        style_cell(ws.cell(r, c), bold=True, fill=P2_FILL, align="center", size=10)
    r += 1

    p2_start = r
    for i, name in enumerate(PHASE2_CANDIDATES, 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, name)
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for c in range(1, 14):
            style_cell(ws.cell(r, c), fill=fill, align="center" if c not in (2, 12, 13) else "left", size=10)
        ws.cell(r, 10, f"=SUM(C{r}:I{r})")
        ws.cell(r, 11, f'=IF(COUNT(C{r}:I{r})>0,AVERAGE(C{r}:I{r}),"")')
        style_cell(ws.cell(r, 10), fill=fill, align="center", bold=True)
        style_cell(ws.cell(r, 11), fill=fill, align="center", bold=True)
        r += 1
    p2_end = r - 1
    add_score_validation(ws, p2_start, p2_end, 3, 9)
    add_score_colors(ws, f"C{p2_start}:I{p2_end}")

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "PHASE 2 CRITERIA GUIDE (simple words)")
    style_cell(ws.cell(r, 1), bold=True, fill=SUBHEADER_FILL, align="center")
    r += 1
    for label, desc in PHASE2_CRITERIA:
        ws.cell(r, 1, label)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(r, 2, desc)
        style_cell(ws.cell(r, 1), bold=True)
        style_cell(ws.cell(r, 2))
        r += 1

    widths = [4, 32, 12, 12, 12, 12, 12, 12, 12, 8, 8, 22, 22]
    set_col_widths(ws, widths)
    ws.freeze_panes = "C5"
    ws.sheet_view.zoomScale = 90


def build_score_sheet(ws, title, candidates, criteria, phase_fill):
    ws.title = title
    ws.merge_cells("A1:F1")
    ws["A1"] = title.upper()
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    headers = ["Candidate"] + [c[0] for c in criteria] + ["Total", "Average", "Overall Notes"]
    ws.append([])
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=phase_fill, align="center")

    for i, name in enumerate(candidates, start=3):
        row_fill = ALT_FILL if i % 2 == 1 else WHITE_FILL
        ws.cell(i, 1, name)
        style_cell(ws.cell(i, 1), fill=row_fill)
        for c in range(2, len(criteria) + 2):
            style_cell(ws.cell(i, c), fill=row_fill, align="center")
        total_col = len(criteria) + 2
        avg_col = len(criteria) + 3
        notes_col = len(criteria) + 4
        first_score = get_column_letter(2)
        last_score = get_column_letter(len(criteria) + 1)
        ws.cell(i, total_col, f"=SUM({first_score}{i}:{last_score}{i})")
        ws.cell(i, avg_col, f'=IF(COUNT({first_score}{i}:{last_score}{i})>0,AVERAGE({first_score}{i}:{last_score}{i}),"")')
        style_cell(ws.cell(i, total_col), fill=row_fill, align="center", bold=True)
        style_cell(ws.cell(i, avg_col), fill=row_fill, align="center", bold=True)
        style_cell(ws.cell(i, notes_col), fill=row_fill)

    last_data = len(candidates) + 2
    add_score_validation(ws, 3, last_data, 2, len(criteria) + 1)
    add_score_colors(ws, f"B3:{get_column_letter(len(criteria)+1)}{last_data}")

    legend_row = last_data + 2
    ws.cell(legend_row, 1, "Criteria guide")
    style_cell(ws.cell(legend_row, 1), bold=True, fill=SUBHEADER_FILL)
    for idx, (label, desc) in enumerate(criteria, start=2):
        ws.cell(legend_row, idx, f"{label}: {desc}")
        style_cell(ws.cell(legend_row, idx), fill=SUBHEADER_FILL)
    ws.cell(legend_row + 1, 1, RATING_SCALE)
    ws.merge_cells(start_row=legend_row + 1, start_column=1, end_row=legend_row + 1, end_column=len(headers))
    style_cell(ws.cell(legend_row + 1, 1), fill=SUBHEADER_FILL)

    widths = [34] + [14] * len(criteria) + [10, 10, 40]
    set_col_widths(ws, widths)
    ws.freeze_panes = "B3"


def build_notes_sheet(ws, title, candidates, criteria):
    ws.title = title
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    row = 3
    for name in candidates:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1, name)
        style_cell(ws.cell(row, 1), bold=True, fill=SUBHEADER_FILL)
        row += 1
        for label, desc in criteria:
            ws.cell(row, 1, label)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row, 2, desc)
            style_cell(ws.cell(row, 1), bold=True)
            style_cell(ws.cell(row, 2))
            row += 2
        row += 1
    set_col_widths(ws, [24, 30, 30, 30])


def build_summary_sheet(wb):
    ws = wb.create_sheet("Summary")
    ws.merge_cells("A1:G1")
    ws["A1"] = "ALL CANDIDATES — QUICK COMPARISON"
    style_cell(ws["A1"], bold=True, fill=HEADER_FILL, align="center", size=14, color=WHITE)

    headers = ["Phase", "Candidate", "Avg Score", "Strengths", "Areas to improve", "Recommend?", "Final notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
        style_cell(ws.cell(2, col), bold=True, fill=SUBHEADER_FILL, align="center")

    row = 3
    for name in PHASE1_CANDIDATES:
        for col, val in enumerate(["Phase 1", name, "", "", "", "", ""], 1):
            ws.cell(row, col, val)
            fill = ALT_FILL if row % 2 == 0 else WHITE_FILL
            style_cell(ws.cell(row, col), fill=fill)
        row += 1
    for name in PHASE2_CANDIDATES:
        for col, val in enumerate(["Phase 2", name, "", "", "", "", ""], 1):
            ws.cell(row, col, val)
            fill = ALT_FILL if row % 2 == 0 else WHITE_FILL
            style_cell(ws.cell(row, col), fill=fill)
        row += 1

    rec_dv = DataValidation(type="list", formula1='"Yes,Maybe,No"', allow_blank=True)
    ws.add_data_validation(rec_dv)
    for r in range(3, row):
        rec_dv.add(ws.cell(r, 6))

    set_col_widths(ws, [10, 34, 12, 28, 28, 14, 30])
    ws.freeze_panes = "A3"


def main():
    wb = Workbook()
    build_dashboard_sheet(wb)

    ws1 = wb.create_sheet("Phase 1 Scores")
    build_score_sheet(ws1, "Session 1 — Phase 1 Scores", PHASE1_CANDIDATES, PHASE1_CRITERIA, P1_FILL)

    ws2 = wb.create_sheet("Phase 2 Scores")
    build_score_sheet(ws2, "Session 2 — Phase 2 Scores", PHASE2_CANDIDATES, PHASE2_CRITERIA, P2_FILL)

    ws3 = wb.create_sheet("Phase 1 Notes")
    build_notes_sheet(ws3, "Phase 1 — Detailed Notes", PHASE1_CANDIDATES, PHASE1_CRITERIA)

    ws4 = wb.create_sheet("Phase 2 Notes")
    build_notes_sheet(ws4, "Phase 2 — Detailed Notes", PHASE2_CANDIDATES, PHASE2_CRITERIA)

    build_summary_sheet(wb)
    wb.save("candidate_evaluation.xlsx")
    print("Created candidate_evaluation.xlsx")


if __name__ == "__main__":
    main()
